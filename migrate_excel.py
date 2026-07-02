"""
Migration script: Import RTS_Game_Design_Heroes.xlsx into the app's format.
One-time run to populate Heroes.xlsx with proper field schema and hero data.
"""
import json
import uuid
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = PROJECT_ROOT / "data" / "RTS_Game_Design_Heroes.xlsx"
OUTPUT_XLSX = PROJECT_ROOT / "data" / "Heroes.xlsx"


def categorize_field(name: str) -> str:
    identity = {"Age", "Temple", "God", "Hero Name", "Role", "Unit Type"}
    combat = {"Health", "Mobility", "Speed", "Melee ATK", "Ranged ATK", "Magic ATK"}
    defense = {"Melee DEF", "Ranged DEF", "Magic DEF", "Armor", "Dodge", "Parry"}
    target = {"Melee Target Type", "Ranged Target Type", "Magical Target Type"}
    range_aoe = {"Range", "Area of Effect"}
    abilities = {"Healing", "Passive", "Abilities", "Ultimate Skill"}
    resource = {"Rage (Required to cast the Ultimate Skill)"}
    economy = {"Cost", "Population"}
    notes = {"Notes"}
    if name in identity: return "Identity"
    if name in combat: return "Combat Stats"
    if name in defense: return "Defense"
    if name in target: return "Target Types"
    if name in range_aoe: return "Range & AoE"
    if name in abilities: return "Abilities"
    if name in resource: return "Resource"
    if name in economy: return "Economy"
    if name in notes: return "Notes"
    return "Custom"


def field_type(name: str) -> str:
    textarea_fields = {"Passive", "Abilities", "Ultimate Skill", "Notes"}
    if name in textarea_fields:
        return "textarea"
    return "text"


def normalize(value):
    return "" if value is None else str(value)


def main():
    if not SOURCE_XLSX.exists():
        print(f"ERROR: Source Excel not found at {SOURCE_XLSX}")
        sys.exit(1)

    print(f"Reading source: {SOURCE_XLSX}")
    wb = openpyxl.load_workbook(SOURCE_XLSX)
    ws = wb["RTS Hero Design"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize(v) for v in rows[0]]

    # Build field definitions from Excel columns + extra fields from HERO_DESIGN.md
    excel_fields = []
    for h in headers:
        if h and h not in {"ID"}:
            excel_fields.append({
                "name": h,
                "category": categorize_field(h),
                "type": field_type(h),
            })

    extra_fields = [
        {"name": "Temple", "category": "Identity", "type": "text"},
        {"name": "God", "category": "Identity", "type": "text"},
        {"name": "Speed", "category": "Combat Stats", "type": "text"},
        {"name": "Range", "category": "Range & AoE", "type": "text"},
        {"name": "Armor", "category": "Defense", "type": "text"},
        {"name": "Cost", "category": "Economy", "type": "text"},
        {"name": "Population", "category": "Economy", "type": "text"},
        {"name": "Area of Effect", "category": "Range & AoE", "type": "text"},
    ]

    existing_names = {f["name"] for f in excel_fields}
    for ef in extra_fields:
        if ef["name"] not in existing_names:
            excel_fields.append(ef)
            existing_names.add(ef["name"])

    print(f"Fields: {len(excel_fields)} total ({len(excel_fields) - len(extra_fields)} from Excel + extras)")

    # Import heroes
    heroes = []
    age_set = set()
    for row in rows[1:]:
        vals = [normalize(v) for v in row]
        hero_name = vals[headers.index("Hero Name")] if "Hero Name" in headers else ""
        if not hero_name:
            continue

        hero = {
            "id": str(uuid.uuid4()),
            "Portrait": "",
            "Extra Images": [],
        }
        for i, h in enumerate(headers):
            if h and h not in {"ID"}:
                hero[h] = vals[i] if i < len(vals) else ""

        # Split "Temple/God" into separate "Temple" and "God" fields
        if "Temple/God" in hero:
            god_value = hero.pop("Temple/God")
            if god_value and god_value != "None":
                hero["God"] = god_value
                hero["Temple"] = f"Temple of {god_value}"
            else:
                hero["God"] = ""
                hero["Temple"] = ""

        # Add empty values for extra fields
        for ef in extra_fields:
            if ef["name"] not in hero:
                hero[ef["name"]] = ""

        age = hero.get("Age", "")
        if age:
            age_set.add(age)
        heroes.append(hero)

    ages = sorted(age_set) if age_set else ["Unassigned"]
    print(f"Heroes: {len(heroes)}")
    print(f"Ages: {ages}")

    # Write output workbook
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    out_wb = openpyxl.Workbook()

    # Fields sheet
    fields_ws = out_wb.create_sheet("Fields")
    fields_ws.append(["name", "category", "type"])
    for f in excel_fields:
        fields_ws.append([f["name"], f["category"], f["type"]])

    # Ages sheet
    ages_ws = out_wb.create_sheet("Ages")
    ages_ws.append(["Age"])
    for age in ages:
        ages_ws.append([age])

    # Heroes sheet
    heroes_ws = out_wb.create_sheet("Heroes")
    heroes_ws.append(["id", "Age", "Hero Name", "Portrait", "Extra Images", "Data JSON"])
    protected = {"id", "Age", "Hero Name", "Portrait", "Extra Images"}
    for hero in heroes:
        payload = {k: v for k, v in hero.items() if k not in protected}
        heroes_ws.append([
            hero.get("id", str(uuid.uuid4())),
            hero.get("Age", ""),
            hero.get("Hero Name", ""),
            hero.get("Portrait", ""),
            json.dumps(hero.get("Extra Images", []), ensure_ascii=False),
            json.dumps(payload, ensure_ascii=False),
        ])

    # Legacy_Export sheet
    flat_headers = [f["name"] for f in excel_fields if f["name"] not in {"ID", "Portrait", "Extra Images"}]
    legacy_ws = out_wb.create_sheet("Legacy_Export")
    legacy_ws.append(flat_headers)
    for hero in heroes:
        legacy_ws.append([hero.get(h, "") for h in flat_headers])

    # Remove default sheet
    if "Sheet" in out_wb.sheetnames:
        out_wb.remove(out_wb["Sheet"])

    out_wb.save(OUTPUT_XLSX)
    print(f"\nSaved: {OUTPUT_XLSX}")
    print(f"  Fields sheet: {len(excel_fields)} fields")
    print(f"  Ages sheet: {len(ages)} ages")
    print(f"  Heroes sheet: {len(heroes)} heroes")
    print("Migration complete!")


if __name__ == "__main__":
    main()
