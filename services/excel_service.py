from typing import Optional, List, Dict, Any
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.database import Database
from models.hero import Hero, HeroImage
from models.age import Age
from models.property import PropertyDefinition, PropertyType


class ExcelService:
    def __init__(self, database: Database):
        self.db = database

    def export_to_excel(self, file_path: Path) -> Path:
        wb = openpyxl.Workbook()

        self._create_heroes_sheet(wb)
        self._create_ages_sheet(wb)
        self._create_properties_sheet(wb)
        self._create_categories_sheet(wb)
        self._create_images_sheet(wb)

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        return file_path

    def _create_heroes_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Heroes")

        prop_defs = self.db.property_definitions
        prop_ids = [p.id for p in prop_defs]
        prop_names = [p.get_display_name() for p in prop_defs]

        headers = ['ID', 'Name', 'Age', 'Description', 'Tags', 'Created', 'Updated'] + prop_names

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        ages_map = {a.id: a.name for a in self.db.ages}

        for row_idx, hero in enumerate(self.db.heroes, 2):
            ws.cell(row=row_idx, column=1, value=hero.id)
            ws.cell(row=row_idx, column=2, value=hero.name)
            ws.cell(row=row_idx, column=3, value=ages_map.get(hero.age_id, ''))
            ws.cell(row=row_idx, column=4, value=hero.description)
            ws.cell(row=row_idx, column=5, value=', '.join(hero.tags))
            ws.cell(row=row_idx, column=6, value=hero.created_at)
            ws.cell(row=row_idx, column=7, value=hero.updated_at)

            for prop_idx, prop_id in enumerate(prop_ids):
                value = hero.properties.get(prop_id, '')
                prop = self.db.get_property(prop_id)
                if prop and prop.property_type == PropertyType.BOOLEAN:
                    value = 'Yes' if value else 'No'
                ws.cell(row=row_idx, column=8 + prop_idx, value=value)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_ages_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Ages")

        headers = ['ID', 'Name', 'Description', 'Order', 'Color']

        header_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, age in enumerate(self.db.ages, 2):
            ws.cell(row=row_idx, column=1, value=age.id)
            ws.cell(row=row_idx, column=2, value=age.name)
            ws.cell(row=row_idx, column=3, value=age.description)
            ws.cell(row=row_idx, column=4, value=age.order)
            ws.cell(row=row_idx, column=5, value=age.color)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_properties_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Properties")

        headers = ['ID', 'Name', 'Display Name', 'Category', 'Type', 'Default Value',
                   'Description', 'Min Value', 'Max Value', 'Enum Values',
                   'Required', 'Visible', 'Order']

        header_fill = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        cat_map = {c.id: c.name for c in self.db.property_categories}

        for row_idx, prop in enumerate(self.db.property_definitions, 2):
            ws.cell(row=row_idx, column=1, value=prop.id)
            ws.cell(row=row_idx, column=2, value=prop.name)
            ws.cell(row=row_idx, column=3, value=prop.display_name)
            ws.cell(row=row_idx, column=4, value=cat_map.get(prop.category_id, ''))
            ws.cell(row=row_idx, column=5, value=prop.property_type.value)
            ws.cell(row=row_idx, column=6, value=str(prop.default_value))
            ws.cell(row=row_idx, column=7, value=prop.description)
            ws.cell(row=row_idx, column=8, value=prop.min_value)
            ws.cell(row=row_idx, column=9, value=prop.max_value)
            ws.cell(row=row_idx, column=10, value=', '.join(prop.enum_values))
            ws.cell(row=row_idx, column=11, value='Yes' if prop.is_required else 'No')
            ws.cell(row=row_idx, column=12, value='Yes' if prop.is_visible else 'No')
            ws.cell(row=row_idx, column=13, value=prop.order)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_categories_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Categories")

        headers = ['ID', 'Name', 'Description', 'Order', 'Collapsed']

        header_fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, cat in enumerate(self.db.property_categories, 2):
            ws.cell(row=row_idx, column=1, value=cat.id)
            ws.cell(row=row_idx, column=2, value=cat.name)
            ws.cell(row=row_idx, column=3, value=cat.description)
            ws.cell(row=row_idx, column=4, value=cat.order)
            ws.cell(row=row_idx, column=5, value='Yes' if cat.is_collapsed else 'No')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_images_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Images")

        headers = ['ID', 'Hero ID', 'Hero Name', 'Image Path', 'Is Primary', 'Caption', 'Order']

        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        hero_names = {h.id: h.name for h in self.db.heroes}
        row_idx = 2

        for hero in self.db.heroes:
            for img in hero.images:
                ws.cell(row=row_idx, column=1, value=img.id)
                ws.cell(row=row_idx, column=2, value=hero.id)
                ws.cell(row=row_idx, column=3, value=hero.name)
                ws.cell(row=row_idx, column=4, value=img.image_path)
                ws.cell(row=row_idx, column=5, value='Yes' if img.is_primary else 'No')
                ws.cell(row=row_idx, column=6, value=img.caption)
                ws.cell(row=row_idx, column=7, value=img.order)
                row_idx += 1

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def import_from_excel(self, file_path: Path) -> Dict[str, int]:
        wb = openpyxl.load_workbook(file_path)
        results = {'heroes': 0, 'ages': 0, 'properties': 0, 'categories': 0, 'images': 0}

        if 'Ages' in wb.sheetnames:
            results['ages'] = self._import_ages(wb['Ages'])

        if 'Categories' in wb.sheetnames:
            results['categories'] = self._import_categories(wb['Categories'])

        if 'Properties' in wb.sheetnames:
            results['properties'] = self._import_properties(wb['Properties'])

        if 'Heroes' in wb.sheetnames:
            results['heroes'] = self._import_heroes(wb['Heroes'])

        if 'Images' in wb.sheetnames:
            results['images'] = self._import_images(wb['Images'])

        return results

    def _import_ages(self, ws) -> int:
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            age_id, name, description, order, color = row[0], row[1], row[2], row[3], row[4]

            existing = self.db.get_age(age_id)
            if existing:
                existing.name = name or ''
                existing.description = description or ''
                existing.order = order or 0
                existing.color = color or '#FFFFFF'
                self.db.update_age(existing)
            else:
                age = Age(
                    id=age_id,
                    name=name or '',
                    description=description or '',
                    order=order or 0,
                    color=color or '#FFFFFF'
                )
                self.db.add_age(age)
            count += 1
        return count

    def _import_categories(self, ws) -> int:
        from models.property import PropertyCategory
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cat_id, name, description, order, collapsed = row[0], row[1], row[2], row[3], row[4]

            existing = self.db.get_category(cat_id)
            if existing:
                existing.name = name or ''
                existing.description = description or ''
                existing.order = order or 0
                existing.is_collapsed = str(collapsed).lower() in ['yes', 'true', '1']
                self.db.update_category(existing)
            else:
                cat = PropertyCategory(
                    id=cat_id,
                    name=name or '',
                    description=description or '',
                    order=order or 0,
                    is_collapsed=str(collapsed).lower() in ['yes', 'true', '1']
                )
                self.db.add_category(cat)
            count += 1
        return count

    def _import_properties(self, ws) -> int:
        count = 0
        cat_map = {c.name: c.id for c in self.db.property_categories}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            prop_id, name, display_name, category, p_type, default_value, description, min_val, max_val, enum_vals, required, visible, order = row[:13]

            existing = self.db.get_property(prop_id)
            enum_list = [e.strip() for e in (enum_vals or '').split(',') if e.strip()] if enum_vals else []

            cat_id = cat_map.get(category)

            if existing:
                existing.name = name or ''
                existing.display_name = display_name or ''
                existing.category_id = cat_id
                existing.property_type = PropertyType(p_type or 'string')
                existing.default_value = default_value
                existing.description = description or ''
                existing.min_value = min_val
                existing.max_value = max_val
                existing.enum_values = enum_list
                existing.is_required = str(required).lower() in ['yes', 'true', '1']
                existing.is_visible = str(visible).lower() in ['yes', 'true', '1']
                existing.order = order or 0
                self.db.update_property(existing)
            else:
                prop = PropertyDefinition(
                    id=prop_id,
                    name=name or '',
                    display_name=display_name or '',
                    category_id=cat_id,
                    property_type=PropertyType(p_type or 'string'),
                    default_value=default_value,
                    description=description or '',
                    min_value=min_val,
                    max_value=max_val,
                    enum_values=enum_list,
                    is_required=str(required).lower() in ['yes', 'true', '1'],
                    is_visible=str(visible).lower() in ['yes', 'true', '1'],
                    order=order or 0
                )
                self.db.add_property(prop)
            count += 1
        return count

    def _import_heroes(self, ws) -> int:
        count = 0
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        prop_indices = {}
        for i, header in enumerate(headers):
            if header and header not in ['ID', 'Name', 'Age', 'Description', 'Tags', 'Created', 'Updated']:
                prop = next((p for p in self.db.property_definitions if p.get_display_name() == header), None)
                if prop:
                    prop_indices[i] = prop.id

        age_map = {a.name: a.id for a in self.db.ages}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            hero_id = row[0]
            name = row[1]
            age_name = row[2]
            description = row[3] or ''
            tags = [t.strip() for t in (row[4] or '').split(',') if t.strip()]
            created = row[5]
            updated = row[6]

            age_id = age_map.get(age_name)

            properties = {}
            for col_idx, prop_id in prop_indices.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None and value != '':
                        prop = self.db.get_property(prop_id)
                        if prop and prop.property_type == PropertyType.INTEGER:
                            try:
                                value = int(value)
                            except (ValueError, TypeError):
                                pass
                        elif prop and prop.property_type == PropertyType.FLOAT:
                            try:
                                value = float(value)
                            except (ValueError, TypeError):
                                pass
                        elif prop and prop.property_type == PropertyType.BOOLEAN:
                            if isinstance(value, str):
                                value = value.lower() in ['yes', 'true', '1']
                            else:
                                value = bool(value)
                        properties[prop_id] = value

            existing = self.db.get_hero(hero_id)
            if existing:
                existing.name = name or ''
                existing.age_id = age_id
                existing.description = description
                existing.tags = tags
                existing.properties = properties
                existing.updated_at = updated or datetime.now().isoformat()
                self.db.update_hero(existing)
            else:
                hero = Hero(
                    id=hero_id,
                    name=name or '',
                    age_id=age_id,
                    description=description,
                    properties=properties,
                    tags=tags,
                    created_at=created or datetime.now().isoformat(),
                    updated_at=updated or datetime.now().isoformat()
                )
                self.db.add_hero(hero)
            count += 1
        return count

    def _import_images(self, ws) -> int:
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            img_id, hero_id, hero_name, img_path, is_primary, caption, order = row[:7]

            hero = self.db.get_hero(hero_id)
            if not hero:
                continue

            existing = None
            for img in hero.images:
                if img.id == img_id:
                    existing = img
                    break

            if existing:
                existing.image_path = img_path or ''
                existing.is_primary = str(is_primary).lower() in ['yes', 'true', '1']
                existing.caption = caption or ''
                existing.order = order or 0
            else:
                new_img = HeroImage(
                    id=img_id,
                    hero_id=hero_id,
                    image_path=img_path or '',
                    is_primary=str(is_primary).lower() in ['yes', 'true', '1'],
                    caption=caption or '',
                    order=order or 0
                )
                hero.images.append(new_img)

            self.db.update_hero(hero)
            count += 1
        return count