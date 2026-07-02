from fastapi import FastAPI, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pathlib import Path
import json
import os
import uuid
import shutil
from typing import Optional, List
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import cloud services (optional - only used when env vars are set)
try:
    from database.supabase_client import SupabaseDB
    from services.cloudinary_service import CloudinaryService
    HAS_CLOUD = True
except ImportError:
    HAS_CLOUD = False

app = FastAPI(title="RTS Game Designer", version="1.0.0")

# Initialize cloud services if credentials are available
supabase_db = SupabaseDB() if HAS_CLOUD else None
cloudinary_svc = CloudinaryService() if HAS_CLOUD else None
IS_CLOUD_MODE = supabase_db and supabase_db.is_cloud if HAS_CLOUD else False

# Static files and templates
APP_ROOT = Path(__file__).resolve().parent
STATIC_DIR = APP_ROOT / "static"
TEMPLATES_DIR = APP_ROOT / "templates"

STATIC_DIR.mkdir(exist_ok=True)
TEMPLATES_DIR.mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

CONFIG_FILE = APP_ROOT / "last_project.json"


def load_config_legacy():
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return {"project_dir": str(APP_ROOT / "data")}


def save_config_legacy(project_dir: Path):
    CONFIG_FILE.write_text(json.dumps({"project_dir": str(project_dir)}, indent=2), encoding="utf-8")


def current_project_dir() -> Path:
    cfg = load_config_legacy()
    return Path(cfg.get("project_dir", APP_ROOT / "data")).expanduser().resolve()


def ensure_project_structure(project_dir: Path):
    project_dir.mkdir(parents=True, exist_ok=True)
    (project_dir / "images").mkdir(parents=True, exist_ok=True)


def workbook_path(project_dir: Path) -> Path:
    return project_dir / "Heroes.xlsx"


def normalize_text(value):
    return "" if value is None else str(value)


def default_project_state():
    return {
        "project_dir": str(current_project_dir()),
        "ages": ["Classical Age", "Medieval Age", "Mythic Age"],
        "categories": ["Identity", "Combat Stats", "Defense", "Target Types", "Range & AoE", "Abilities", "Resource", "Economy", "Notes", "Custom"],
        "roles": [],
        "gods": [],
        "kingdoms": ["Humans", "Kingdom of Ores", "Kingdom of Magic", "Kingdom of Evil", "Kingdom of the Ancients", "Kingdom of the Beasts", "Kingdom of the Divines"],
        "kingdoms_data": {
            "Humans": {"description": "Versatile and balanced. Masters of adaptation and technology.", "forefathers": "First Men", "established": "Classical Age", "races": "Humans", "history": "The First Men built the earliest civilizations. Over ages they mastered warfare, diplomacy, and technology. Though individually weaker than other races, their unity and innovation allowed them to thrive."},
            "Kingdom of Ores": {"description": "Mountain-dwelling master craftsmen and heavy infantry.", "forefathers": "Durin the Deathless", "established": "Classical Age", "races": "Dwarves", "history": "Forged in the heart of mountains, the dwarves were created by the god Moradin from stone and fire. They built labyrinthine halls beneath peaks, mastered metallurgy and rune-magic."},
            "Kingdom of Magic": {"description": "Arcane-focused faction wielding powerful spellcraft.", "forefathers": "Archmage Aelindor", "established": "Medieval Age", "races": "Elves, Wizards, Gnomes", "history": "When the veil between worlds thinned, the Elves emerged from the Feywilds. Aelindor taught mortals to harness mana. The Gnomes brought alchemical ingenuity. They built the Crystal Spires."},
            "Kingdom of Evil": {"description": "Practitioners of dark magic, necromancy, and forbidden arts.", "forefathers": "Lich Lord Vorthak", "established": "Medieval Age", "races": "Necromancers, Witches, Goblins", "history": "Vorthak was once a noble wizard who sought immortality. His experiments with death magic corrupted him. From his Black Citadel, he raised the first undead legions."},
            "Kingdom of the Ancients": {"description": "Primitive but powerful beings bonded with primordial beasts.", "forefathers": "Great Mother Sauria", "established": "Classical Age", "races": "Silurian, Dinosaurs", "history": "Before civilization, the Silurian roamed jungles alongside dinosaurs. Great Mother Sauria was the first to bond with a T-Rex through blood rituals. They reject magic and technology."},
            "Kingdom of the Beasts": {"description": "Monstrous creatures of immense physical power.", "forefathers": "Baal the Horned King", "established": "Classical Age", "races": "Minotaurs, Trolls, Balrogs, Dragons, Serpents", "history": "Baal was the first Minotaur to unite monster tribes beneath volcanic peaks. Trolls, Balrogs, and Dragons pledged fealty to his strength. They value dominance above all."},
            "Kingdom of the Divines": {"description": "Divine beings of holy power and celestial might.", "forefathers": "The Creator Athenia", "established": "Classical Age", "races": "Gods, Angels, Demigods, Oracles, Archangels", "history": "Athenia, the Creator, breathed life into the first angels from starlight. They built the Celestial Citadel above the clouds. The Divines represent absolute order and judgment against darkness."},
        },
        "fields": [],
        "heroes": [],
    }


def slugify(value: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in normalize_text(value))
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_") or "hero"


def project_file(project_dir: Path) -> Path:
    return workbook_path(project_dir)


def ensure_workbook(project_dir: Path):
    ensure_project_structure(project_dir)
    if project_file(project_dir).exists():
        return
    seed_from_legacy_columns(project_dir)


def field_lookup(fields):
    return {field["name"]: field for field in fields}


def seed_from_legacy_columns(project_dir: Path):
    state = default_project_state()
    state["project_dir"] = str(project_dir)
    samples = [
        {
            "id": str(uuid.uuid4()),
            "Age": "Mythic Age",
            "Kingdom": "Kingdom of the Divines",
            "Hero Name": "Paladin",
            "Portrait": "",
            "Extra Images": [],
            "Temple": "Temple of Athenia",
            "God": "Athenia",
            "Role": "Heavy Tank",
            "Health": "Extremely High",
            "Mobility": "Low-Moderate",
            "Melee ATK": "Very High",
            "Ranged ATK": "",
            "Magic ATK": "Low (Holy attack)",
            "Melee DEF": "Extremely High",
            "Ranged DEF": "Moderately High",
            "Magic DEF": "Moderate",
            "Healing": "",
            "Passive": "Divine Retribution (Counterattack every hit)",
            "Abilities": "Heavy frontline tank",
            "Ultimate Skill": "Bulwark of Athenia",
            "Notes": "Strong vs Archers, weaker vs Mages",
            "Rage (Required to cast the Ultimate Skill)": "1000",
            "Dodge": "High",
            "Parry": "Very High",
        },
        {
            "id": str(uuid.uuid4()),
            "Age": "Mythic Age",
            "Kingdom": "Kingdom of Magic",
            "Hero Name": "Sword Master",
            "Portrait": "",
            "Extra Images": [],
            "Temple": "Temple of Barathor",
            "God": "Barathor",
            "Role": "Duelist",
            "Health": "Moderate",
            "Mobility": "High",
            "Melee ATK": "Extremely High",
            "Ranged ATK": "",
            "Magic ATK": "Moderate (later levels)",
            "Melee DEF": "Moderate",
            "Ranged DEF": "Moderate",
            "Magic DEF": "Moderate",
            "Healing": "",
            "Passive": "Battle Focus",
            "Abilities": "Whirlwind Blades",
            "Ultimate Skill": "Barathor's Wrath",
            "Notes": "No lifesteal",
            "Rage (Required to cast the Ultimate Skill)": "800",
            "Dodge": "Very High",
            "Parry": "Moderate",
        },
    ]
    state["heroes"] = samples
    write_state(project_dir, state)


def import_existing_workbook(project_dir: Path):
    path = project_file(project_dir)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if {"Heroes", "Fields", "Ages"}.issubset(set(wb.sheetnames)):
        return read_state(project_dir)

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        seed_from_legacy_columns(project_dir)
        return read_state(project_dir)

    headers = [normalize_text(v) for v in rows[0] if normalize_text(v)]
    base_fields = [
        {"name": "ID", "category": "System", "type": "text"},
        {"name": "Portrait", "category": "Media", "type": "image"},
        {"name": "Extra Images", "category": "Media", "type": "images"},
    ]
    fields = base_fields + [
        {
            "name": header,
            "category": categorize_field(header),
            "type": "textarea" if header in {"Passive", "Abilities", "Ultimate Skill", "Notes"} else "text",
        }
        for header in headers
        if header not in {"Portrait", "Extra Images", "ID"}
    ]
    # Ensure Kingdom field always exists
    if not any(f.get("name") == "Kingdom" for f in fields):
        fields.insert(0, {"name": "Kingdom", "category": "Identity", "type": "text"})
    heroes = []
    age_values = []
    for row in rows[1:]:
        values = {headers[i]: normalize_text(row[i]) if i < len(row) else "" for i in range(len(headers))}
        if not any(values.values()):
            continue
        hero = {"id": str(uuid.uuid4()), "Portrait": "", "Extra Images": [], "Kingdom": ""}
        hero.update(values)
        if not hero.get("Kingdom"):
            hero["Kingdom"] = "Humans"
        if values.get("Age"):
            age_values.append(values["Age"])
        heroes.append(hero)
    state = {
        "project_dir": str(project_dir),
        "ages": unique_preserve(age_values) or ["Unassigned"],
        "fields": fields,
        "heroes": heroes,
    }
    write_state(project_dir, state)
    return state


def categorize_field(name: str) -> str:
    combat = {
        "Health", "Mobility", "Melee ATK", "Ranged ATK", "Magic ATK",
        "Melee DEF", "Ranged DEF", "Magic DEF", "Range", "Armor",
        "Cost", "Population", "Rage (Required to cast the Ultimate Skill)",
        "Dodge", "Parry", "Area of Effect",
    }
    identity = {"Age", "Hero Name", "Temple", "God", "Role", "Unit Type", "Goddess", "Temple", "Kingdom"}
    abilities = {"Passive", "Abilities", "Ultimate Skill", "Special Skills", "Passive Abilities", "Ultimate Ability"}
    media = {"Portrait", "Extra Images"}
    support = {"Healing", "Healing Ability"}
    if name in identity:
        return "Identity"
    if name in combat:
        return "Combat"
    if name in abilities:
        return "Abilities"
    if name in media:
        return "Media"
    if name in support:
        return "Support"
    if name == "Notes":
        return "Notes"
    return "Custom"


def unique_preserve(items):
    seen = set()
    ordered = []
    for item in items:
        text = normalize_text(item)
        if text and text not in seen:
            seen.add(text)
            ordered.append(text)
    return ordered


def read_state(project_dir: Path):
    ensure_project_structure(project_dir)
    ensure_workbook(project_dir)
    path = project_file(project_dir)
    from openpyxl import load_workbook
    wb = load_workbook(path)

    if not {"Heroes", "Fields", "Ages"}.issubset(set(wb.sheetnames)):
        return import_existing_workbook(project_dir)

    fields_ws = wb["Fields"]
    ages_ws = wb["Ages"]
    heroes_ws = wb["Heroes"]

    fields = []
    for row in fields_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        fields.append({
            "name": normalize_text(row[0]),
            "category": normalize_text(row[1]) or "Custom",
            "type": normalize_text(row[2]) or "text",
        })
    # Ensure Kingdom field always exists
    if not any(f.get("name") == "Kingdom" for f in fields):
        fields.insert(0, {"name": "Kingdom", "category": "Identity", "type": "text"})

    ages = [normalize_text(row[0]) for row in ages_ws.iter_rows(min_row=2, values_only=True) if normalize_text(row[0])]

    # Infer fields from heroes if fields list is empty (backward compatibility)
    if not fields and heroes:
        inferred = set()
        for hero in heroes:
            inferred.update(hero.keys())
        fields = [
            {"name": "ID", "category": "System", "type": "text"},
            {"name": "Portrait", "category": "Media", "type": "image"},
            {"name": "Extra Images", "category": "Media", "type": "images"},
        ]
        for name in sorted(inferred):
            if name not in {"id", "Portrait", "Extra Images", "Kingdom"}:
                fields.append({
                    "name": name,
                    "category": categorize_field(name),
                    "type": "textarea" if name in {"Passive", "Abilities", "Ultimate Skill", "Notes"} else "text",
                })
        # Ensure Kingdom is included
        if not any(f.get("name") == "Kingdom" for f in fields):
            fields.insert(0, {"name": "Kingdom", "category": "Identity", "type": "text"})

    # Read kingdoms
    kingdoms = []
    kingdoms_data = {}
    if "Kingdoms" in wb.sheetnames:
        kingdoms_ws = wb["Kingdoms"]
        for row in kingdoms_ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                kingdoms.append(normalize_text(row[0]))
                kingdoms_data[normalize_text(row[0])] = {
                    "description": normalize_text(row[1]) if len(row) > 1 else "",
                    "forefathers": normalize_text(row[2]) if len(row) > 2 else "",
                    "established": normalize_text(row[3]) if len(row) > 3 else "",
                    "races": normalize_text(row[4]) if len(row) > 4 else "",
                    "history": normalize_text(row[5]) if len(row) > 5 else "",
                }
    if not kingdoms:
        kingdoms = ["Humans", "Kingdom of Ores", "Kingdom of Magic", "Kingdom of Evil", "Kingdom of the Ancients", "Kingdom of the Beasts", "Kingdom of the Divines"]

    # Read heroes: prefer per-kingdom sheets (K_*), fallback to Heroes sheet
    heroes = []
    kingdom_sheets = [s for s in wb.sheetnames if s.startswith("K_")]
    if kingdom_sheets:
        for sheet_name in kingdom_sheets:
            ws = wb[sheet_name]
            kingdom_name = sheet_name[2:].replace("_", " ")
            # Try to find the actual kingdom name from the sheet header if available
            if ws.max_row >= 1:
                first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if first_row and len(first_row[0]) >= 7 and first_row[0][6]:
                    header = first_row[0]
                    # Check if there's a Kingdom column header
                    pass
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                hero = {
                    "id": normalize_text(row[0]),
                    "Age": normalize_text(row[1]),
                    "Hero Name": normalize_text(row[2]),
                    "Portrait": normalize_text(row[3]),
                    "Extra Images": json.loads(row[4]) if row[4] else [],
                }
                data = json.loads(row[5]) if row[5] else {}
                for key, value in data.items():
                    hero[key] = value
                if not hero.get("Kingdom"):
                    hero["Kingdom"] = kingdom_name
                heroes.append(hero)
    else:
        # Fallback: read from Heroes sheet (backward compatibility)
        for row in heroes_ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            hero = {
                "id": normalize_text(row[0]),
                "Age": normalize_text(row[1]),
                "Hero Name": normalize_text(row[2]),
                "Portrait": normalize_text(row[3]),
                "Extra Images": json.loads(row[4]) if row[4] else [],
            }
            data = json.loads(row[5]) if row[5] else {}
            for key, value in data.items():
                hero[key] = value
            if not hero.get("Kingdom"):
                hero["Kingdom"] = kingdoms[0] if kingdoms else "Humans"
            heroes.append(hero)

    categories = []
    if "Categories" in wb.sheetnames:
        cat_ws = wb["Categories"]
        categories = [normalize_text(row[0]) for row in cat_ws.iter_rows(min_row=2, values_only=True) if normalize_text(row[0])]
    if not categories:
        categories = sorted(set(f.get("category", "Custom") for f in fields)) or ["Custom"]

    roles = []
    if "Roles" in wb.sheetnames:
        roles_ws = wb["Roles"]
        roles = [normalize_text(row[0]) for row in roles_ws.iter_rows(min_row=2, values_only=True) if normalize_text(row[0])]
    if not roles:
        roles = sorted(set(h.get("Role", "") for h in heroes if h.get("Role"))) or []

    gods = []
    if "Gods" in wb.sheetnames:
        gods_ws = wb["Gods"]
        gods = [normalize_text(row[0]) for row in gods_ws.iter_rows(min_row=2, values_only=True) if normalize_text(row[0])]
    if not gods:
        gods = sorted(set(h.get("God", "") for h in heroes if h.get("God") and h.get("God") != "None")) or []

    return {"project_dir": str(project_dir), "fields": fields or [], "ages": ages or ["Unassigned"], "categories": categories, "roles": roles, "gods": gods, "kingdoms": kingdoms, "kingdoms_data": kingdoms_data, "heroes": heroes}


def write_state(project_dir: Path, state):
    ensure_project_structure(project_dir)
    from openpyxl import Workbook
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    fields_ws = wb.create_sheet("Fields")
    fields_ws.append(["name", "category", "type"])
    for field in state["fields"]:
        fields_ws.append([field["name"], field.get("category", "Custom"), field.get("type", "text")])

    ages_ws = wb.create_sheet("Ages")
    ages_ws.append(["Age"])
    for age in state["ages"]:
        ages_ws.append([age])

    categories_ws = wb.create_sheet("Categories")
    categories_ws.append(["Category"])
    for cat in state.get("categories", []):
        categories_ws.append([cat])

    roles_ws = wb.create_sheet("Roles")
    roles_ws.append(["Role"])
    for role in state.get("roles", []):
        roles_ws.append([role])

    gods_ws = wb.create_sheet("Gods")
    gods_ws.append(["God"])
    for god in state.get("gods", []):
        gods_ws.append([god])

    # Write Kingdoms sheet
    kingdoms_ws = wb.create_sheet("Kingdoms")
    kingdoms_ws.append(["Kingdom", "Description", "Forefathers", "Established", "Races", "History"])
    kingdoms_data = state.get("kingdoms_data", {})
    for kingdom in state.get("kingdoms", []):
        data = kingdoms_data.get(kingdom, {})
        kingdoms_ws.append([
            kingdom,
            data.get("description", ""),
            data.get("forefathers", ""),
            data.get("established", ""),
            data.get("races", ""),
            data.get("history", ""),
        ])

    # Write per-kingdom hero sheets
    kingdoms = state.get("kingdoms", ["Humans"])
    protected = {"id", "Age", "Hero Name", "Portrait", "Extra Images"}
    for kingdom in kingdoms:
        sheet_name = "K_" + kingdom.replace(" ", "_")
        # Ensure valid Excel sheet name (max 31 chars)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append(["id", "Age", "Hero Name", "Portrait", "Extra Images", "Data JSON"])
        for hero in state["heroes"]:
            if hero.get("Kingdom") == kingdom:
                payload = {k: v for k, v in hero.items() if k not in protected}
                ws.append([
                    hero.get("id", str(uuid.uuid4())),
                    hero.get("Age", ""),
                    hero.get("Hero Name", ""),
                    hero.get("Portrait", ""),
                    json.dumps(hero.get("Extra Images", []), ensure_ascii=False),
                    json.dumps(payload, ensure_ascii=False),
                ])

    # Also write a unified Heroes sheet for backward compatibility
    heroes_ws = wb.create_sheet("Heroes")
    heroes_ws.append(["id", "Age", "Hero Name", "Portrait", "Extra Images", "Data JSON"])
    for hero in state["heroes"]:
        payload = {k: v for k, v in hero.items() if k not in protected}
        heroes_ws.append([
            hero.get("id", str(uuid.uuid4())),
            hero.get("Age", ""),
            hero.get("Hero Name", ""),
            hero.get("Portrait", ""),
            json.dumps(hero.get("Extra Images", []), ensure_ascii=False),
            json.dumps(payload, ensure_ascii=False),
        ])

    legacy_ws = wb.create_sheet("Legacy_Export")
    flat_headers = [field["name"] for field in state["fields"] if field["name"] not in {"ID", "Portrait", "Extra Images"}]
    legacy_ws.append(flat_headers)
    for hero in state["heroes"]:
        legacy_ws.append([hero.get(header, "") for header in flat_headers])

    wb.save(project_file(project_dir))
    save_config_legacy(project_dir)


def current_state():
    return read_state(current_project_dir())


def save_image(file_storage, project_dir: Path):
    if not file_storage or not file_storage.filename:
        return ""
    ext = Path(file_storage.filename).suffix.lower() or ".png"
    safe_name = f"{uuid.uuid4().hex}{ext}"
    output = project_dir / "images" / safe_name
    file_storage.save(output)
    return safe_name


def hero_display_type(hero):
    return hero.get("Unit Type") or hero.get("Role") or "Unknown"


INDEX_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Game Character Designer</title>
  <style>
    :root {
      --bg: #10141c;
      --panel: #181e29;
      --panel-2: #20293a;
      --accent: #6ea8fe;
      --accent-2: #8ef0c8;
      --danger: #f07178;
      --text: #ebf0f8;
      --muted: #9ba8be;
      --border: #2b3648;
      --highlight: rgba(110, 168, 254, 0.16);
      --diff: rgba(240, 113, 120, 0.16);
      --shadow: 0 12px 30px rgba(0, 0, 0, 0.3);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: Inter, Segoe UI, Roboto, Arial, sans-serif;
      background: linear-gradient(180deg, #0c1017 0%, #111827 100%);
      color: var(--text);
      height: 100vh;
      overflow: hidden;
    }
    button, input, select, textarea {
      font: inherit;
      color: inherit;
    }
    .app {
      display: grid;
      grid-template-rows: 56px 1fr;
      height: 100vh;
    }
    .topbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 0 16px;
      background: rgba(24, 30, 41, 0.92);
      border-bottom: 1px solid var(--border);
      backdrop-filter: blur(10px);
    }
    .brand {
      display: flex;
      align-items: center;
      gap: 12px;
      font-weight: 700;
      letter-spacing: 0.02em;
    }
    .brand-badge {
      width: 28px;
      height: 28px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      background: linear-gradient(135deg, var(--accent), var(--accent-2));
      color: #09111f;
      font-weight: 900;
    }
    .toolbar, .toolbar-left {
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
    }
    .layout {
      display: grid;
      grid-template-columns: 280px 1fr 280px;
      gap: 12px;
      padding: 12px;
      min-height: 0;
    }
    .panel {
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 16px;
      box-shadow: var(--shadow);
      min-height: 0;
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }
    .panel-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 12px 14px;
      border-bottom: 1px solid var(--border);
      background: rgba(255,255,255,0.02);
    }
    .panel-body {
      padding: 12px;
      overflow: auto;
    }
    .sidebar-list, .hero-list {
      display: flex;
      flex-direction: column;
      gap: 8px;
    }
    .age-row, .hero-card, .nav-chip, .compare-chip {
      border: 1px solid var(--border);
      background: var(--panel-2);
      border-radius: 12px;
      padding: 10px;
      cursor: pointer;
      transition: 0.15s ease;
    }
    .age-row:hover, .hero-card:hover, .nav-chip:hover, .compare-chip:hover {
      border-color: var(--accent);
      transform: translateY(-1px);
    }
    .age-row.active, .hero-card.active, .nav-chip.active {
      border-color: var(--accent);
      background: var(--highlight);
    }
    .hero-card-title {
      font-weight: 700;
      margin-bottom: 4px;
    }
    .muted { color: var(--muted); }
    .small { font-size: 12px; }
    .controls { display: grid; gap: 8px; }
    .input, .select, .textarea {
      width: 100%;
      background: #111723;
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 10px 12px;
      outline: none;
    }
    .input:focus, .select:focus, .textarea:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(110,168,254,0.16);
    }
    .textarea { min-height: 90px; resize: vertical; }
    .btn {
      border: 1px solid var(--border);
      background: #182131;
      border-radius: 10px;
      padding: 9px 12px;
      cursor: pointer;
    }
    .btn:hover { border-color: var(--accent); }
    .btn.primary {
      background: linear-gradient(135deg, #3f7ae0, #5da8ff);
      border-color: transparent;
      color: white;
    }
    .btn.danger {
      background: rgba(240,113,120,0.12);
      border-color: rgba(240,113,120,0.35);
      color: #ffc6cb;
    }
    .row { display: flex; gap: 8px; align-items: center; }
    .row.wrap { flex-wrap: wrap; }
    .grow { flex: 1; }
    .editor {
      display: grid;
      grid-template-rows: auto 1fr;
      min-height: 0;
      flex: 1;
    }
    .hero-header {
      display: grid;
      grid-template-columns: 220px 1fr;
      gap: 16px;
      padding: 14px;
      border-bottom: 1px solid var(--border);
    }
    .portrait-box {
      border: 1px dashed var(--border);
      border-radius: 16px;
      min-height: 250px;
      display: flex;
      align-items: center;
      justify-content: center;
      background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.01));
      overflow: hidden;
      position: relative;
    }
    .portrait-box img {
      width: 100%;
      height: 100%;
      object-fit: cover;
    }
    .meta-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(180px, 1fr));
      gap: 12px;
    }
    .field-card {
      border: 1px solid var(--border);
      border-radius: 14px;
      margin-bottom: 12px;
      overflow: hidden;
      background: rgba(255,255,255,0.02);
    }
    .field-card summary {
      list-style: none;
      cursor: pointer;
      padding: 12px 14px;
      font-weight: 700;
      background: rgba(255,255,255,0.03);
      border-bottom: 1px solid var(--border);
    }
    .field-card summary::-webkit-details-marker { display: none; }
    .field-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(220px, 1fr));
      gap: 12px;
      padding: 12px;
    }
    .field {
      display: grid;
      gap: 6px;
      align-content: start;
    }
    .field.wide { grid-column: 1 / -1; }
    .badge {
      padding: 4px 8px;
      border-radius: 999px;
      background: rgba(142,240,200,0.14);
      border: 1px solid rgba(142,240,200,0.25);
      color: #b9ffe6;
      font-size: 12px;
      white-space: nowrap;
    }
    .compare-view, .manage-view { display: none; height: 100%; }
    .compare-view.active, .manage-view.active { display: block; }
    .compare-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }
    .compare-table th, .compare-table td {
      border: 1px solid var(--border);
      padding: 10px;
      text-align: left;
      vertical-align: top;
    }
    .compare-table th {
      position: sticky;
      top: 0;
      background: #1c2534;
      z-index: 1;
    }
    .compare-table tr.diff td { background: var(--diff); }
    .thumb-strip {
      display: flex;
      gap: 8px;
      overflow: auto;
      padding-top: 8px;
    }
    .thumb-strip img {
      width: 64px;
      height: 64px;
      object-fit: cover;
      border-radius: 10px;
      border: 1px solid var(--border);
    }
    .section-title {
      font-size: 13px;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.08em;
      margin-bottom: 8px;
    }
    .status {
      padding: 8px 12px;
      border-radius: 999px;
      background: rgba(110,168,254,0.1);
      border: 1px solid rgba(110,168,254,0.2);
      color: #d7e6ff;
      font-size: 12px;
    }
    .empty {
      padding: 24px;
      text-align: center;
      color: var(--muted);
      border: 1px dashed var(--border);
      border-radius: 16px;
    }
    .separator { height: 1px; background: var(--border); margin: 10px 0; }
    @media (max-width: 1200px) {
      .layout { grid-template-columns: 260px 1fr; }
      .right-panel { display: none; }
    }
    @media (max-width: 900px) {
      body { overflow: auto; height: auto; }
      .app, .layout { height: auto; }
      .layout { grid-template-columns: 1fr; }
      .hero-header { grid-template-columns: 1fr; }
      .meta-grid, .field-grid { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <div class="app">
    <div class="topbar">
      <div class="brand">
        <div class="brand-badge">G</div>
        <div>
          <div>Game Character Designer</div>
          <div class="muted small">Offline hero editor with Excel-backed storage</div>
        </div>
      </div>
      <div class="toolbar-left">
        <button class="btn" onclick="switchMode('editor')">Editor</button>
        <button class="btn" onclick="switchMode('compare')">Compare Heroes</button>
        <button class="btn" onclick="switchMode('manage')">Manage Schema</button>
        <button class="btn" onclick="switchMode('docs')">Design Docs</button>
      </div>
      <div class="toolbar">
        <div class="status" id="saveStatus">Ready</div>
        <button class="btn" onclick="promptProjectPath()">Project Folder</button>
        <button class="btn" onclick="downloadWorkbook()">Download Workbook</button>
        <button class="btn primary" onclick="saveProject(true)">Save</button>
      </div>
    </div>

    <div class="layout">
      <div class="panel left-panel">
        <div class="panel-header"><strong>Ages & Heroes</strong><button class="btn" onclick="createHero()">+ Hero</button></div>
        <div class="panel-body">
          <div class="controls">
            <label class="small muted">Kingdom</label>
            <select id="kingdomFilter" class="select" onchange="selectKingdom(this.value)"></select>
            <input id="searchBox" class="input" placeholder="Search hero..." oninput="renderAll()" />
            <div class="row">
              <select id="typeFilter" class="select grow" onchange="renderAll()"></select>
              <button class="btn" onclick="createRole()" title="Add Role">+</button>
              <button class="btn" onclick="manageRole()" title="Rename/Delete selected role">...</button>
            </div>
          </div>
          <div class="separator"></div>
          <div class="row wrap" style="margin-bottom:8px;">
            <button class="btn" onclick="createAge()">+ Age</button>
            <button class="btn" onclick="renameAge()">Rename</button>
            <button class="btn danger" onclick="deleteAge()">Delete</button>
          </div>
          <div id="ageList" class="sidebar-list"></div>
          <div class="separator"></div>
          <div id="heroList" class="hero-list"></div>
        </div>
      </div>

      <div class="panel">
        <div id="editorView" class="editor">
          <div class="panel-header">
            <div class="row wrap">
              <strong id="editorTitle">Hero Editor</strong>
              <span class="badge" id="heroTypeBadge">No selection</span>
            </div>
            <div class="row wrap">
              <button class="btn" onclick="prevHero()">Previous Hero</button>
              <button class="btn" onclick="nextHero()">Next Hero</button>
              <button class="btn" onclick="duplicateHero()">Duplicate</button>
              <button class="btn danger" onclick="deleteHero()">Delete</button>
            </div>
          </div>
          <div id="editorContent" class="panel-body"></div>
        </div>

        <div id="compareView" class="compare-view panel-body">
          <div class="row wrap" style="margin-bottom: 12px;">
            <strong>Compare Heroes</strong>
            <span class="muted small">Select two or more heroes from the right sidebar</span>
            <button class="btn primary" onclick="runCompare()">Refresh Comparison</button>
          </div>
          <div id="compareContent" class="empty">Choose heroes to compare.</div>
        </div>

        <div id="manageView" class="manage-view panel-body">
          <div class="row wrap" style="justify-content: space-between; margin-bottom: 12px;">
            <div>
              <strong>Manage Schema</strong>
              <div class="muted small">Manage categories and fields.</div>
            </div>
          </div>

          <details class="field-card" open>
            <summary>Feature Categories</summary>
            <div class="panel-body">
              <div class="row wrap" style="margin-bottom: 8px;">
                <button class="btn primary" onclick="createCategory()">+ Category</button>
              </div>
              <div id="categoryListContent"></div>
            </div>
          </details>

          <details class="field-card" open style="margin-top: 12px;">
            <summary>Roles / Unit Types</summary>
            <div class="panel-body">
              <div class="row wrap" style="margin-bottom: 8px;">
                <button class="btn primary" onclick="createRole()">+ Role</button>
              </div>
              <div id="roleListContent"></div>
            </div>
          </details>

          <details class="field-card" open style="margin-top: 12px;">
            <summary>Kingdoms</summary>
            <div class="panel-body">
              <div class="row wrap" style="margin-bottom: 8px;">
                <button class="btn primary" onclick="createKingdom()">+ Kingdom</button>
              </div>
              <div id="kingdomListContent"></div>
            </div>
          </details>

          <details class="field-card" open style="margin-top: 12px;">
            <summary>Fields</summary>
            <div class="panel-body">
              <div class="row wrap" style="margin-bottom: 8px;">
                <button class="btn primary" onclick="addField()">+ Field</button>
              </div>
              <div id="schemaContent"></div>
            </div>
          </details>
        </div>

        <div id="docsView" class="manage-view panel-body">
          <div class="row wrap" style="justify-content: space-between; margin-bottom: 12px; align-items: center;">
            <div class="row wrap" style="gap: 12px;">
              <strong>Design Documents</strong>
              <span class="muted small">For developers & designers — edit and track changes</span>
            </div>
            <div class="row wrap">
              <select id="docSelect" class="select" style="min-width: 200px;" onchange="loadDoc()"></select>
              <button class="btn" id="docEditToggle" onclick="toggleDocEditMode()">Edit</button>
              <button class="btn" onclick="createNewDoc()">+ New</button>
              <button class="btn primary" id="docSaveBtn" onclick="saveDoc()" style="display:none;">Save & Commit</button>
            </div>
          </div>
          <div class="row" style="gap: 12px; align-items: stretch; height: calc(100% - 60px);">
            <!-- Editor (hidden in view mode) -->
            <div id="docEditorPanel" style="flex: 1; display: none; flex-direction: column; gap: 8px;">
              <label class="small muted">Markdown Editor</label>
              <textarea id="docEditor" class="textarea" style="flex: 1; min-height: 400px; font-family: 'Consolas', 'Monaco', monospace; font-size: 13px; line-height: 1.5;" placeholder="# Title\n\nWrite your design document here..." oninput="renderDocPreview()"></textarea>
            </div>
            <!-- Viewer -->
            <div id="docViewerPanel" style="flex: 1; display: flex; flex-direction: column; gap: 8px;">
              <div class="row wrap" style="justify-content: space-between; align-items: center;">
                <label class="small muted" id="docTitleLabel">Document Viewer</label>
                <span class="badge" id="docVersionBadge">Latest</span>
              </div>
              <div id="docPreview" class="panel-body" style="flex: 1; overflow: auto; background: var(--panel-2); border-radius: 10px; padding: 24px; font-size: 14px; line-height: 1.6;"></div>
            </div>
          </div>
        </div>
      </div>

      <div class="panel right-panel">
        <div class="panel-header"><strong id="rightPanelTitle">Navigation & Tools</strong></div>
        <div class="panel-body">
          <div id="rightPanelStandard">
            <div class="section-title">Compare Selection</div>
            <div id="comparePicker" class="sidebar-list"></div>
            <div class="separator"></div>
            <div class="section-title">Feature Categories</div>
            <div id="categoryNav" class="sidebar-list"></div>
          </div>
          <div id="rightPanelDocs" style="display:none;">
            <div class="section-title">Version History</div>
            <div id="docHistory" class="sidebar-list" style="font-size: 12px;">
              <div class="empty">Select a document to see its history.</div>
            </div>
            <div class="separator"></div>
            <div class="section-title">Document Info</div>
            <div id="docInfo" class="muted small" style="padding: 8px 0;">
              Changes are automatically tracked via Git. Every save creates a version you can review.
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <script>
    const state = {
      project: null,
      selectedKingdom: null,
      selectedAge: null,
      selectedHeroId: null,
      mode: 'editor',
      compareIds: [],
      autosaveTimer: null,
      periodicTimer: null,
      lastSaveTime: null,
    };

    const protectedFields = new Set(['Age', 'Hero Name', 'Portrait', 'Extra Images']);

    async function api(url, options = {}) {
      const response = await fetch(url, {
        headers: { ...(options.body instanceof FormData ? {} : { 'Content-Type': 'application/json' }) },
        ...options,
      });
      if (!response.ok) {
        const data = await response.json().catch(() => ({ error: 'Request failed' }));
        throw new Error(data.error || 'Request failed');
      }
      return response.json();
    }

    function setStatus(text) {
      document.getElementById('saveStatus').textContent = text;
    }

    function imgUrl(file) {
      return file ? `/images/${encodeURIComponent(file)}` : 'https://placehold.co/400x520?text=Portrait';
    }

    async function init() {
      await loadProject();
      switchMode('editor');
      state.periodicTimer = setInterval(periodicSave, 600000);
    }

    // Document / Design Docs functions
    let docEditMode = false;
    let currentDocHistory = [];

    async function loadDocList() {
      const docs = await api('/api/docs/list');
      const select = document.getElementById('docSelect');
      const current = select.value || docs[0]?.filename || '';
      select.innerHTML = docs.map(d => `<option value="${escapeAttr(d.filename)}" ${d.filename === current ? 'selected' : ''}>${escapeHtml(d.filename.replace('.md', ''))}</option>`).join('');
      if (select.value) loadDoc();
    }

    async function loadDoc() {
      const filename = document.getElementById('docSelect').value;
      if (!filename) return;
      const doc = await api(`/api/docs/read?filename=${encodeURIComponent(filename)}`);
      document.getElementById('docEditor').value = doc.content;
      document.getElementById('docTitleLabel').textContent = filename.replace('.md', '');
      document.getElementById('docVersionBadge').textContent = 'Latest';
      renderDocPreview();
      await loadDocHistory(filename);
    }

    function toggleDocEditMode() {
      docEditMode = !docEditMode;
      const editorPanel = document.getElementById('docEditorPanel');
      const viewerPanel = document.getElementById('docViewerPanel');
      const saveBtn = document.getElementById('docSaveBtn');
      const toggleBtn = document.getElementById('docEditToggle');
      
      if (docEditMode) {
        editorPanel.style.display = 'flex';
        viewerPanel.style.flex = '1';
        saveBtn.style.display = 'inline-block';
        toggleBtn.textContent = 'View';
        toggleBtn.classList.add('primary');
        renderDocPreview();
      } else {
        editorPanel.style.display = 'none';
        viewerPanel.style.flex = '1';
        saveBtn.style.display = 'none';
        toggleBtn.textContent = 'Edit';
        toggleBtn.classList.remove('primary');
      }
    }

    function renderDocPreview() {
      const markdown = document.getElementById('docEditor').value;
      const preview = document.getElementById('docPreview');
      
      // Enhanced markdown-to-HTML renderer
      let html = markdown
        .replace(/^# (.*$)/gim, '<h1 style="font-size: 24px; border-bottom: 2px solid var(--accent); padding-bottom: 8px; margin-bottom: 16px; color: var(--text);">$1</h1>')
        .replace(/^## (.*$)/gim, '<h2 style="font-size: 18px; margin-top: 24px; margin-bottom: 12px; color: var(--accent);">$1</h2>')
        .replace(/^### (.*$)/gim, '<h3 style="font-size: 15px; margin-top: 18px; margin-bottom: 8px; color: var(--muted);">$1</h3>')
        .replace(/^---+/gim, '<hr style="border: none; border-top: 1px solid var(--border); margin: 20px 0;">')
        .replace(/\*\*(.*?)\*\*/g, '<strong style="color: var(--accent-2);">$1</strong>')
        .replace(/\*(.*?)\*/g, '<em>$1</em>')
        .replace(/`([^`]+)`/g, '<code style="background: rgba(110,168,254,0.1); padding: 2px 6px; border-radius: 4px; font-family: monospace; font-size: 12px;">$1</code>')
        .replace(/^\* (.*$)/gim, '<li style="margin-left: 20px; margin-bottom: 4px;">$1</li>')
        .replace(/^\d+\.\s(.*$)/gim, '<li style="margin-left: 20px; margin-bottom: 4px; list-style-type: decimal;">$1</li>')
        .replace(/^\|(.+)\|$/gim, (match, p1) => {
          const cells = p1.split('|').map(c => `<td style="border: 1px solid var(--border); padding: 8px; text-align: left;">${c.trim()}</td>`).join('');
          return `<tr>${cells}</tr>`;
        })
        .replace(/\[([^\]]+)\]\(([^)]+)\)/g, '<a href="$2" target="_blank" style="color: var(--accent); text-decoration: none;">$1</a>')
        .replace(/\n/g, '<br>');
      
      // Wrap consecutive li in ul
      html = html.replace(/(<li[^>]*>.*<\/li>)(?:<br>)+/g, '<ul style="margin-bottom: 12px;">$1</ul>');
      // Wrap consecutive tr in table
      html = html.replace(/(<tr>.*<\/tr>)(?:<br>)+/g, '<table style="width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 13px;">$1</table>');
      
      // Style tables
      html = html.replace(/<table/g, '<table style="width: 100%; border-collapse: collapse; margin: 12px 0;"');
      html = html.replace(/<td>/g, '<td style="border: 1px solid var(--border); padding: 8px;">');
      
      preview.innerHTML = html;
    }

    async function loadDocHistory(filename) {
      try {
        const data = await api(`/api/docs/history?filename=${encodeURIComponent(filename)}`);
        currentDocHistory = data.history || [];
        const container = document.getElementById('docHistory');
        
        if (currentDocHistory.length === 0) {
          container.innerHTML = '<div class="empty">No history yet. Save the document to create the first version.</div>';
          return;
        }
        
        container.innerHTML = currentDocHistory.map((h, i) => `
          <div class="nav-chip ${i === 0 ? 'active' : ''}" onclick="loadDocVersion('${escapeAttr(h.hash)}', ${i})" style="padding: 8px 10px; font-size: 11px;">
            <div style="font-weight: 600; margin-bottom: 2px;">${escapeHtml(h.message)}</div>
            <div class="muted" style="font-size: 10px;">${escapeHtml(h.hash)} • ${escapeHtml(h.date)}</div>
          </div>
        `).join('');
      } catch (e) {
        document.getElementById('docHistory').innerHTML = '<div class="empty">History not available.</div>';
      }
    }

    async function loadDocVersion(commitHash, index) {
      const filename = document.getElementById('docSelect').value;
      if (!filename || !commitHash) return;
      
      setStatus('Loading version...');
      try {
        const data = await api(`/api/docs/version?filename=${encodeURIComponent(filename)}&commit=${encodeURIComponent(commitHash)}`);
        document.getElementById('docEditor').value = data.content;
        document.getElementById('docVersionBadge').textContent = `Version: ${commitHash}`;
        renderDocPreview();
        
        // Highlight selected version
        const chips = document.getElementById('docHistory').querySelectorAll('.nav-chip');
        chips.forEach((c, i) => c.classList.toggle('active', i === index));
        
        setStatus('Loaded version ' + commitHash);
      } catch (e) {
        setStatus('Failed to load version');
      }
    }

    async function saveDoc() {
      const filename = document.getElementById('docSelect').value;
      const content = document.getElementById('docEditor').value;
      if (!filename) return;
      const message = prompt('Describe this change (commit message):', `Updated ${filename.replace('.md', '')}`);
      if (message === null) return;
      setStatus('Saving doc...');
      await api('/api/docs/save', {
        method: 'POST',
        body: JSON.stringify({ filename, content, message })
      });
      setStatus('Saved & committed');
      // Reload history
      await loadDocHistory(filename);
      document.getElementById('docVersionBadge').textContent = 'Latest';
    }

    async function createNewDoc() {
      const name = prompt('New document name (e.g., COMBAT.md):', 'NEW_DESIGN.md');
      if (!name) return;
      const content = `# ${name.replace('.md', '')}\n\nAdd your design notes here.`;
      await api('/api/docs/save', {
        method: 'POST',
        body: JSON.stringify({ filename: name, content, message: `Created ${name}` })
      });
      await loadDocList();
      document.getElementById('docSelect').value = name;
      loadDoc();
    }

    async function loadProject() {
      state.project = await api('/api/project');
      if (!state.project.ages.length) state.project.ages = ['Unassigned'];
      if (!state.project.kingdoms || !state.project.kingdoms.length) {
        state.project.kingdoms = ['Humans', 'Kingdom of Ores', 'Kingdom of Magic', 'Kingdom of Evil', 'Kingdom of the Ancients', 'Kingdom of the Beasts', 'Kingdom of the Divines'];
      }
      if (!state.project.kingdoms_data) {
        state.project.kingdoms_data = {};
      }
      state.selectedKingdom = state.selectedKingdom && state.project.kingdoms.includes(state.selectedKingdom)
        ? state.selectedKingdom
        : state.project.kingdoms[0];
      state.selectedAge = state.selectedAge && state.project.ages.includes(state.selectedAge)
        ? state.selectedAge
        : state.project.ages[0];
      const heroes = filteredHeroes(false);
      if (!state.selectedHeroId || !state.project.heroes.some(h => h.id === state.selectedHeroId)) {
        state.selectedHeroId = heroes[0]?.id || state.project.heroes[0]?.id || null;
      }
      syncCompareSelection();
      renderAll();
    }

    function switchMode(mode) {
      state.mode = mode;
      document.getElementById('editorView').style.display = mode === 'editor' ? 'grid' : 'none';
      document.getElementById('compareView').classList.toggle('active', mode === 'compare');
      document.getElementById('manageView').classList.toggle('active', mode === 'manage');
      document.getElementById('docsView').classList.toggle('active', mode === 'docs');
      
      // Update right panel based on mode
      const rightStandard = document.getElementById('rightPanelStandard');
      const rightDocs = document.getElementById('rightPanelDocs');
      const rightTitle = document.getElementById('rightPanelTitle');
      
      if (mode === 'docs') {
        rightStandard.style.display = 'none';
        rightDocs.style.display = 'block';
        rightTitle.textContent = 'Version History';
        loadDocList();
      } else {
        rightStandard.style.display = 'block';
        rightDocs.style.display = 'none';
        rightTitle.textContent = 'Navigation & Tools';
      }
      
      renderAll();
    }

    function getFieldDefs() {
      return state.project?.fields || [];
    }

    function getHeroById(id) {
      return state.project.heroes.find(hero => hero.id === id);
    }

    function fieldGroups() {
      const groups = {};
      for (const field of getFieldDefs()) {
        const category = field.category || 'Custom';
        groups[category] ||= [];
        groups[category].push(field);
      }
      return groups;
    }

    function filteredHeroes(byAge = true) {
      const search = document.getElementById('searchBox')?.value?.trim().toLowerCase() || '';
      return state.project.heroes.filter(hero => {
        const inKingdom = !state.selectedKingdom || hero.Kingdom === state.selectedKingdom;
        const inAge = !byAge || !state.selectedAge || hero.Age === state.selectedAge;
        const haystack = [hero['Hero Name'], hero.God, hero.Temple, hero.Role, hero['Unit Type'], hero.Passive, hero.Abilities]
          .filter(Boolean).join(' ').toLowerCase();
        const matchesSearch = !search || haystack.includes(search);
        return inKingdom && inAge && matchesSearch;
      });
    }

    function populateKingdomFilter() {
      const select = document.getElementById('kingdomFilter');
      if (!select || !state.project) return;
      const current = state.selectedKingdom || state.project.kingdoms[0] || '';
      select.innerHTML = state.project.kingdoms.map(k => `<option value="${escapeAttr(k)}" ${k === current ? 'selected' : ''}>${escapeHtml(k)}</option>`).join('');
    }

    function selectKingdom(kingdom) {
      state.selectedKingdom = kingdom;
      const heroes = filteredHeroes(true);
      if (heroes.length && !heroes.some(h => h.id === state.selectedHeroId)) {
        state.selectedHeroId = heroes[0]?.id || null;
      }
      renderAll();
    }

    function populateTypeFilter() {
      const select = document.getElementById('typeFilter');
      const current = select.value || 'All Types';
      const roles = state.project?.roles || [];
      const types = ['All Types', ...roles];
      select.innerHTML = types.map(type => `<option ${type === current ? 'selected' : ''}>${escapeHtml(type)}</option>`).join('');
    }

    function renderAll() {
      if (!state.project) return;
      populateKingdomFilter();
      populateTypeFilter();
      renderAges();
      renderHeroes();
      renderEditor();
      renderCategoryNav();
      renderComparePicker();
      renderSchema();
      renderCategories();
      renderRoles();
      renderKingdoms();
      if (state.mode === 'compare') runCompare();
    }

    function renderAges() {
      const ageList = document.getElementById('ageList');
      ageList.innerHTML = state.project.ages.map(age => {
        const count = state.project.heroes.filter(hero => hero.Age === age && (!state.selectedKingdom || hero.Kingdom === state.selectedKingdom)).length;
        return `<div class="age-row ${age === state.selectedAge ? 'active' : ''}" onclick="selectAge(${js(age)})">
          <div class="row" style="justify-content: space-between;">
            <strong>${escapeHtml(age)}</strong>
            <span class="badge">${count}</span>
          </div>
        </div>`;
      }).join('');
    }

    function renderHeroes() {
      const heroList = document.getElementById('heroList');
      const heroes = filteredHeroes(true);
      if (!heroes.length) {
        heroList.innerHTML = '<div class="empty">No heroes match the current filters.</div>';
        return;
      }
      heroList.innerHTML = heroes.map(hero => {
        const type = hero['Unit Type'] || hero.Role || 'Unknown';
        const kingdomData = state.project?.kingdoms_data?.[hero.Kingdom] || {};
        const kingdomLabel = kingdomData.description 
          ? `${escapeHtml(hero.Kingdom || '')}: ${escapeHtml(kingdomData.description)}`
          : escapeHtml(hero.Kingdom || '');
        return `<div class="hero-card ${hero.id === state.selectedHeroId ? 'active' : ''}" onclick="selectHero(${js(hero.id)})">
          <div class="hero-card-title">${escapeHtml(hero['Hero Name'] || 'Unnamed Hero')}</div>
          <div class="small muted">${escapeHtml(type)}</div>
          <div class="small muted" title="${escapeAttr(kingdomData.description || '')}">${kingdomLabel} &middot; ${escapeHtml(hero.God || hero.Age || '')}</div>
        </div>`;
      }).join('');
    }

    function renderEditor() {
      const container = document.getElementById('editorContent');
      const hero = getHeroById(state.selectedHeroId);
      document.getElementById('editorTitle').textContent = hero ? (hero['Hero Name'] || 'Unnamed Hero') : 'Hero Editor';
      document.getElementById('heroTypeBadge').textContent = hero ? (hero['Unit Type'] || hero.Role || 'Unknown') : 'No selection';
      if (!hero) {
        container.innerHTML = '<div class="empty">Create or select a hero to start editing.</div>';
        return;
      }

      const groups = fieldGroups();
      const coreFields = ['Hero Name', 'Kingdom', 'Age', 'Temple', 'God', 'Role', 'Unit Type'];
      const metaGrid = coreFields.filter(name => getFieldDefs().some(f => f.name === name)).map(name => renderInput(hero, getFieldDefs().find(f => f.name === name), true)).join('');
      const categoryMarkup = Object.entries(groups).map(([category, fields]) => {
        const inner = fields
          .filter(field => !['ID', 'Portrait', 'Extra Images'].includes(field.name))
          .map(field => renderInput(hero, field))
          .join('');
        if (!inner) return '';
        return `<details class="field-card" open id="cat-${slug(category)}">
          <summary>${escapeHtml(category)}</summary>
          <div class="field-grid">${inner}</div>
        </details>`;
      }).join('');

      container.innerHTML = `
        <div class="hero-header">
          <div>
            <div class="portrait-box">
              <img src="${imgUrl(hero.Portrait)}" alt="Portrait preview" />
            </div>
            <div class="controls" style="margin-top: 10px;">
              <label class="btn">Upload Portrait<input type="file" accept="image/*" style="display:none" onchange="uploadPortrait(event)" /></label>
              <label class="btn">Add Images<input type="file" accept="image/*" multiple style="display:none" onchange="uploadExtraImages(event)" /></label>
            </div>
            <div class="thumb-strip">${(hero['Extra Images'] || []).map(file => `<img src="${imgUrl(file)}" alt="Extra image" />`).join('')}</div>
          </div>
          <div>
            <div class="meta-grid">${metaGrid}</div>
            <div class="separator"></div>
            <div class="row wrap">
              <button class="btn" onclick="renameHero()">Rename</button>
              <button class="btn" onclick="moveHeroToAge()">Move to Age</button>
              <button class="btn" onclick="toggleCompareCurrent()">${state.compareIds.includes(hero.id) ? 'Remove from Compare' : 'Add to Compare'}</button>
            </div>
          </div>
        </div>
        <div>${categoryMarkup}</div>
      `;
    }

    function renderInput(hero, field, compact = false) {
      if (!field) return '';
      const name = field.name;
      const value = hero[name] ?? '';
      const label = `<label class="small muted">${escapeHtml(name)}</label>`;
      const wide = field.type === 'textarea' ? 'wide' : '';
      if (name === 'Kingdom') {
        const kingdoms = state.project.kingdoms || [];
        return `<div class="field ${wide}">${label}<div class="row" style="gap:6px">
          <select class="select" style="flex:1" onchange="updateField(${js(hero.id)}, ${js(name)}, this.value)">
            ${kingdoms.map(k => `<option value="${escapeAttr(k)}" ${k === value ? 'selected' : ''}>${escapeHtml(k)}</option>`).join('')}
          </select>
          <button class="btn btn--small btn--flat" onclick="manageKingdoms()" title="Manage Kingdoms">...</button>
        </div></div>`;
      }
      if (name === 'Age') {
        return `<div class="field ${wide}">${label}<select class="select" onchange="updateField(${js(hero.id)}, ${js(name)}, this.value)">
          ${state.project.ages.map(age => `<option value="${escapeAttr(age)}" ${age === value ? 'selected' : ''}>${escapeHtml(age)}</option>`).join('')}
        </select></div>`;
      }
      if (name === 'Temple') {
        return `<div class="field ${wide}">${label}<input class="input" value="${escapeAttr(value)}" oninput="updateField(${js(hero.id)}, ${js(name)}, this.value)" /></div>`;
      }
      if (name === 'God') {
        const gods = state.project.gods || [];
        return `<div class="field ${wide}">${label}<div class="row" style="gap:6px">
          <select class="select" style="flex:1" onchange="updateField(${js(hero.id)}, ${js(name)}, this.value)">
            <option value="">None</option>
            ${gods.map(g => `<option value="${escapeAttr(g)}" ${g === value ? 'selected' : ''}>${escapeHtml(g)}</option>`).join('')}
          </select>
          <button class="btn btn--small btn--flat" onclick="manageGods()" title="Manage Gods">...</button>
        </div></div>`;
      }
      if (field.type === 'textarea') {
        return `<div class="field ${wide}">${label}<textarea class="textarea" oninput="updateField(${js(hero.id)}, ${js(name)}, this.value)">${escapeHtml(value)}</textarea></div>`;
      }
      if (field.type === 'images') {
        return `<div class="field wide">${label}<div class="thumb-strip">${(value || []).map(file => `<img src="${imgUrl(file)}" alt="Image" />`).join('')}</div></div>`;
      }
      if (field.type === 'image') {
        return `<div class="field ${wide}">${label}<input class="input" value="${escapeAttr(value)}" disabled /></div>`;
      }
      return `<div class="field ${compact ? '' : wide}">${label}<input class="input" value="${escapeAttr(value)}" oninput="updateField(${js(hero.id)}, ${js(name)}, this.value)" /></div>`;
    }

    function renderCategoryNav() {
      const nav = document.getElementById('categoryNav');
      const groups = Object.keys(fieldGroups());
      nav.innerHTML = groups.map(category => `<div class="nav-chip" onclick="scrollToCategory(${js(category)})">${escapeHtml(category)}</div>`).join('');
    }

    function renderComparePicker() {
      const picker = document.getElementById('comparePicker');
      const selected = state.project.heroes.filter(hero => state.compareIds.includes(hero.id));
      if (!selected.length) {
        picker.innerHTML = '<div class="empty">No heroes selected.</div>';
        return;
      }
      picker.innerHTML = selected.map(hero => {
        return `<div class="compare-chip active" onclick="toggleCompare(${js(hero.id)})">
          <div class="row" style="justify-content: space-between; gap: 6px;">
            <span>${escapeHtml(hero['Hero Name'] || 'Unnamed Hero')}</span>
            <span class="badge">${escapeHtml(hero.Kingdom || hero.Age || 'Age')}</span>
          </div>
        </div>`;
      }).join('');
    }

    function renderSchema() {
      const container = document.getElementById('schemaContent');
      const fields = getFieldDefs();
      const categories = state.project?.categories || [];
      container.innerHTML = fields.map(field => `
        <div class="field-card">
          <div class="field-grid">
            <div class="field">
              <label class="small muted">Field Name</label>
              <input class="input" value="${escapeAttr(field.name)}" onchange="renameField(${js(field.name)}, this.value, null, null)" ${protectedFields.has(field.name) ? 'disabled' : ''} />
            </div>
            <div class="field">
              <label class="small muted">Category</label>
              <select class="select" onchange="renameField(${js(field.name)}, null, this.value, null)">
                ${categories.map(cat => `<option value="${escapeAttr(cat)}" ${cat === (field.category || 'Custom') ? 'selected' : ''}>${escapeHtml(cat)}</option>`).join('')}
              </select>
            </div>
            <div class="field">
              <label class="small muted">Type</label>
              <select class="select" onchange="renameField(${js(field.name)}, null, null, this.value)">
                ${['text', 'textarea', 'image', 'images'].map(type => `<option value="${type}" ${type === field.type ? 'selected' : ''}>${type}</option>`).join('')}
              </select>
            </div>
            <div class="field" style="align-content:end;">
              <button class="btn danger" onclick="deleteField(${js(field.name)})" ${protectedFields.has(field.name) ? 'disabled' : ''}>Delete Field</button>
            </div>
          </div>
        </div>`).join('');
    }

    function renderCategories() {
      const container = document.getElementById('categoryListContent');
      const categories = state.project?.categories || [];
      const fields = getFieldDefs();
      container.innerHTML = categories.map(cat => {
        const count = fields.filter(f => f.category === cat).length;
        return `<div class="field-card">
          <div class="field-grid">
            <div class="field">
              <label class="small muted">Feature Categories</label>
              <input class="input" value="${escapeAttr(cat)}" onchange="renameCategory(${js(cat)}, this.value)" />
            </div>
            <div class="field">
              <label class="small muted">Fields in category</label>
              <div class="input" style="opacity:0.6;cursor:default;">${count} field${count !== 1 ? 's' : ''}</div>
            </div>
            <div class="field" style="align-content:end;">
              <button class="btn danger" onclick="deleteCategory(${js(cat)})">Delete</button>
            </div>
          </div>
        </div>`;
      }).join('');
    }

    async function createCategory() {
      const name = prompt('Category name:', 'New Category');
      if (!name) return;
      const data = await api('/api/category', { method: 'POST', body: JSON.stringify({ name }) });
      state.project.categories = data.categories;
      renderAll();
      queueAutosave();
    }

    async function renameCategory(oldName, newName) {
      if (!newName || newName === oldName) return;
      const data = await api(`/api/category/${encodeURIComponent(oldName)}`, { method: 'PUT', body: JSON.stringify({ name: newName }) });
      state.project.categories = data.categories;
      state.project.fields = data.fields;
      renderAll();
      queueAutosave();
    }

    async function deleteCategory(name) {
      const fields = getFieldDefs().filter(f => f.category === name);
      const msg = fields.length > 0
        ? `Delete category "${name}"? ${fields.length} field(s) will be moved to "Custom".`
        : `Delete category "${name}"?`;
      if (!confirm(msg)) return;
      const data = await api(`/api/category/${encodeURIComponent(name)}`, { method: 'DELETE' });
      state.project.categories = data.categories;
      state.project.fields = data.fields;
      renderAll();
      queueAutosave();
    }

    function renderRoles() {
      const container = document.getElementById('roleListContent');
      const roles = state.project?.roles || [];
      const heroes = state.project?.heroes || [];
      container.innerHTML = roles.map(role => {
        const count = heroes.filter(h => h.Role === role).length;
        return `<div class="field-card">
          <div class="field-grid">
            <div class="field">
              <label class="small muted">Role Name</label>
              <input class="input" value="${escapeAttr(role)}" onchange="renameRole(${js(role)}, this.value)" />
            </div>
            <div class="field">
              <label class="small muted">Heroes with this role</label>
              <div class="input" style="opacity:0.6;cursor:default;">${count} hero${count !== 1 ? 'es' : ''}</div>
            </div>
            <div class="field" style="align-content:end;">
              <button class="btn danger" onclick="deleteRole(${js(role)})">Delete</button>
            </div>
          </div>
        </div>`;
      }).join('');
    }

    async function createRole() {
      const name = prompt('Role name:', 'New Role');
      if (!name) return;
      const data = await api('/api/role', { method: 'POST', body: JSON.stringify({ name }) });
      state.project.roles = data.roles;
      renderAll();
      queueAutosave();
    }

    async function renameRole(oldName, newName) {
      if (!newName || newName === oldName) return;
      const data = await api(`/api/role/${encodeURIComponent(oldName)}`, { method: 'PUT', body: JSON.stringify({ name: newName }) });
      state.project.roles = data.roles;
      state.project.heroes.forEach(h => { if (h.Role === oldName) h.Role = newName; });
      renderAll();
      queueAutosave();
    }

    async function deleteRole(name) {
      const heroes = (state.project?.heroes || []).filter(h => h.Role === name);
      const msg = heroes.length > 0
        ? `Delete role "${name}"? ${heroes.length} hero(s) will have their Role cleared.`
        : `Delete role "${name}"?`;
      if (!confirm(msg)) return;
      const data = await api(`/api/role/${encodeURIComponent(name)}`, { method: 'DELETE' });
      state.project.roles = data.roles;
      state.project.heroes.forEach(h => { if (h.Role === name) h.Role = ''; });
      renderAll();
      queueAutosave();
    }

    function manageRole() {
      const select = document.getElementById('typeFilter');
      const current = select.value;
      if (!current || current === 'All Types') {
        alert('Select a specific role first, then click ... to rename or delete it.');
        return;
      }
      const action = prompt(`Role: "${current}"\n\nType a new name to rename, or leave blank to delete:`, current);
      if (action === null) return;
      if (action === '') {
        deleteRole(current);
      } else if (action !== current) {
        renameRole(current, action);
      }
    }

    function renderKingdoms() {
      const container = document.getElementById('kingdomListContent');
      const kingdoms = state.project?.kingdoms || [];
      const kingdomsData = state.project?.kingdoms_data || {};
      const heroes = state.project?.heroes || [];
      container.innerHTML = kingdoms.map(kingdom => {
        const count = heroes.filter(h => h.Kingdom === kingdom).length;
        const data = kingdomsData[kingdom] || {};
        return `<div class="field-card">
          <div class="field-grid">
            <div class="field">
              <label class="small muted">Kingdom Name</label>
              <input class="input" value="${escapeAttr(kingdom)}" onchange="renameKingdom(${js(kingdom)}, this.value)" />
            </div>
            <div class="field">
              <label class="small muted">Heroes</label>
              <div class="input" style="opacity:0.6;cursor:default;">${count} hero${count !== 1 ? 'es' : ''}</div>
            </div>
            <div class="field" style="align-content:end;">
              <button class="btn" onclick="editKingdomDetails(${js(kingdom)})">Edit Details</button>
              <button class="btn danger" onclick="deleteKingdom(${js(kingdom)})">Delete</button>
            </div>
          </div>
          <div class="small muted" style="padding: 0 14px 10px;">
            ${escapeHtml(data.description || '')}
            ${data.forefathers ? `<br><strong>Forefathers:</strong> ${escapeHtml(data.forefathers)}` : ''}
            ${data.races ? ` | <strong>Races:</strong> ${escapeHtml(data.races)}` : ''}
          </div>
        </div>`;
      }).join('');
    }

    async function createKingdom() {
      const name = prompt('Kingdom name:', 'New Kingdom');
      if (!name) return;
      const data = await api('/api/kingdom', { method: 'POST', body: JSON.stringify({ name }) });
      state.project.kingdoms = data.kingdoms;
      state.project.kingdoms_data = data.kingdoms_data;
      renderAll();
      queueAutosave();
    }

    async function renameKingdom(oldName, newName) {
      if (!newName || newName === oldName) return;
      const data = state.project.kingdoms_data?.[oldName] || {};
      const payload = {
        name: newName,
        description: data.description || '',
        forefathers: data.forefathers || '',
        established: data.established || '',
        races: data.races || '',
        history: data.history || '',
      };
      const result = await api(`/api/kingdom/${encodeURIComponent(oldName)}`, { method: 'PUT', body: JSON.stringify(payload) });
      state.project.kingdoms = result.kingdoms;
      state.project.kingdoms_data = result.kingdoms_data;
      state.project.heroes.forEach(h => { if (h.Kingdom === oldName) h.Kingdom = newName; });
      renderAll();
      queueAutosave();
    }

    async function editKingdomDetails(kingdom) {
      const data = state.project.kingdoms_data?.[kingdom] || {};
      const description = prompt(`Description for ${kingdom}:`, data.description || '');
      if (description === null) return;
      const forefathers = prompt('Forefathers:', data.forefathers || '');
      if (forefathers === null) return;
      const established = prompt('Established (Age):', data.established || '');
      if (established === null) return;
      const races = prompt('Races:', data.races || '');
      if (races === null) return;
      const history = prompt('History (short):', data.history || '');
      if (history === null) return;
      
      const payload = {
        name: kingdom,
        description,
        forefathers,
        established,
        races,
        history,
      };
      const result = await api(`/api/kingdom/${encodeURIComponent(kingdom)}`, { method: 'PUT', body: JSON.stringify(payload) });
      state.project.kingdoms_data = result.kingdoms_data;
      renderAll();
      queueAutosave();
    }

    async function deleteKingdom(name) {
      const heroes = (state.project?.heroes || []).filter(h => h.Kingdom === name);
      const msg = heroes.length > 0
        ? `Delete kingdom "${name}"? ${heroes.length} hero(s) will be moved to "${state.project.kingdoms[0] || 'Humans'}".`
        : `Delete kingdom "${name}"?`;
      if (!confirm(msg)) return;
      const data = await api(`/api/kingdom/${encodeURIComponent(name)}`, { method: 'DELETE' });
      state.project.kingdoms = data.kingdoms;
      state.project.kingdoms_data = data.kingdoms_data;
      state.project.heroes.forEach(h => { if (h.Kingdom === name) h.Kingdom = state.project.kingdoms[0] || 'Humans'; });
      renderAll();
      queueAutosave();
    }

    function manageKingdoms() {
      const kingdoms = state.project.kingdoms || [];
      const list = kingdoms.length > 0 ? kingdoms.join(', ') : '(none)';
      const action = prompt(`Kingdoms:\n${list}\n\nOptions:\n• Type a kingdom name to add it\n• To rename: old name = new name\n• To delete: type just the name (will remove)\n\nEnter kingdom name:`, '');
      if (action === null || action === '') return;
      if (action.includes('=')) {
        const [oldName, newName] = action.split('=').map(s => s.trim());
        if (!oldName || !newName) return;
        renameKingdom(oldName, newName);
      } else if (kingdoms.includes(action)) {
        const rename = prompt(`Delete kingdom "${action}"? (type yes to confirm)`, '');
        if (rename === 'yes') deleteKingdom(action);
      } else {
        createKingdom(action);
      }
    }

    function manageGods() {
      const gods = state.project.gods || [];
      const list = gods.length > 0 ? gods.join(', ') : '(none)';
      const action = prompt(`Gods / Temples:\n${list}\n\nOptions:\n• Type a god name to add it\n• To rename: old name = new name\n• To delete: type just the name (will remove)\n\nEnter god name:`, '');
      if (action === null || action === '') return;
      if (action.includes('=')) {
        const [oldName, newName] = action.split('=').map(s => s.trim());
        if (!oldName || !newName) return;
        renameGod(oldName, newName);
      } else if (gods.includes(action)) {
        const rename = prompt(`Delete god "${action}"? (type yes to confirm)`, '');
        if (rename === 'yes') deleteGod(action);
      } else {
        createGod(action);
      }
    }

    async function createGod(name) {
      const data = await api('/api/god', { method: 'POST', body: JSON.stringify({ name }) });
      state.project.gods = data.gods;
      renderAll();
      queueAutosave();
    }

    async function renameGod(oldName, newName) {
      if (!newName || newName === oldName) return;
      const data = await api(`/api/god/${encodeURIComponent(oldName)}`, { method: 'PUT', body: JSON.stringify({ name: newName }) });
      state.project.gods = data.gods;
      state.project.heroes.forEach(h => { if (h.God === oldName) h.God = newName; });
      renderAll();
      queueAutosave();
    }

    async function deleteGod(name) {
      const heroes = (state.project?.heroes || []).filter(h => h.God === name);
      const msg = heroes.length > 0
        ? `Delete god "${name}"? ${heroes.length} hero(s) will have their God cleared.`
        : `Delete god "${name}"?`;
      if (!confirm(msg)) return;
      const data = await api(`/api/god/${encodeURIComponent(name)}`, { method: 'DELETE' });
      state.project.gods = data.gods;
      state.project.heroes.forEach(h => { if (h.God === name) h.God = ''; });
      renderAll();
      queueAutosave();
    }

    async function runCompare() {
      const content = document.getElementById('compareContent');
      if (state.compareIds.length < 2) {
        content.className = 'empty';
        content.innerHTML = 'Choose two or more heroes to compare.';
        return;
      }
      const data = await api('/api/compare', { method: 'POST', body: JSON.stringify({ hero_ids: state.compareIds }) });
      const heroHeader = data.heroes.map(hero => `
        <th>
          <div>${escapeHtml(hero['Hero Name'] || 'Unnamed Hero')}</div>
          <div class="small muted">${escapeHtml(hero.Kingdom || '')} &middot; ${escapeHtml(hero.Age || '')} &middot; ${escapeHtml(hero['Unit Type'] || hero.Role || 'Unknown')}</div>
          <div style="margin-top:8px;"><img src="${imgUrl(hero.Portrait)}" alt="Portrait" style="width:120px;height:150px;object-fit:cover;border-radius:12px;border:1px solid var(--border);" /></div>
        </th>`).join('');
      const rows = data.rows.map(row => `
        <tr class="${row.different ? 'diff' : ''}">
          <td><strong>${escapeHtml(row.field)}</strong></td>
          ${row.values.map(value => `<td>${escapeHtml(value || '&mdash;')}</td>`).join('')}
        </tr>`).join('');
      content.className = '';
      content.innerHTML = `<table class="compare-table"><thead><tr><th>Field</th>${heroHeader}</tr></thead><tbody>${rows}</tbody></table>`;
    }

    function selectAge(age) {
      state.selectedAge = age;
      const heroes = filteredHeroes(true);
      if (heroes.length && !heroes.some(h => h.id === state.selectedHeroId)) state.selectedHeroId = heroes[0].id;
      renderAll();
    }

    function selectHero(id) {
      state.selectedHeroId = id;
      renderAll();
    }

    function syncCompareSelection() {
      state.compareIds = state.compareIds.filter(id => state.project.heroes.some(hero => hero.id === id));
    }

    async function updateField(heroId, fieldName, value) {
      const hero = getHeroById(heroId);
      hero[fieldName] = value;
      if (fieldName === 'Age') state.selectedAge = value;
      renderHeroes();
      document.getElementById('editorTitle').textContent = hero['Hero Name'] || 'Unnamed Hero';
      document.getElementById('heroTypeBadge').textContent = hero['Unit Type'] || hero.Role || 'Unknown';
      queueAutosave();
    }

    function queueAutosave() {
      clearTimeout(state.autosaveTimer);
      setStatus('Unsaved changes');
      state.autosaveTimer = setTimeout(() => saveProject(false), 500);
    }

    async function saveProject(manual) {
      if (!state.project) return;
      setStatus(manual ? 'Saving...' : 'Auto-saving...');
      state.project.project_dir ||= '';
      await api('/api/save', { method: 'POST', body: JSON.stringify(state.project) });
      state.lastSaveTime = new Date();
      const timeStr = state.lastSaveTime.toLocaleTimeString();
      setStatus(manual ? 'Saved at ' + timeStr : 'Auto-saved at ' + timeStr);
    }

    async function periodicSave() {
      if (!state.project) return;
      await api('/api/save', { method: 'POST', body: JSON.stringify(state.project) });
      state.lastSaveTime = new Date();
      const timeStr = state.lastSaveTime.toLocaleTimeString();
      setStatus('Periodic save at ' + timeStr);
    }

    async function createHero() {
      const name = prompt('Hero name:', 'New Hero');
      if (name === null) return;
      const hero = await api('/api/hero', { method: 'POST', body: JSON.stringify({ 'Hero Name': name, Age: state.selectedAge, Kingdom: state.selectedKingdom }) });
      state.project.heroes.push(hero);
      state.selectedHeroId = hero.id;
      renderAll();
      queueAutosave();
    }

    async function duplicateHero() {
      if (!state.selectedHeroId) return;
      const hero = await api(`/api/hero/${state.selectedHeroId}/duplicate`, { method: 'POST' });
      state.project.heroes.push(hero);
      state.selectedHeroId = hero.id;
      renderAll();
    }

    async function deleteHero() {
      const hero = getHeroById(state.selectedHeroId);
      if (!hero || !confirm(`Delete ${hero['Hero Name']}?`)) return;
      await api(`/api/hero/${hero.id}`, { method: 'DELETE' });
      state.project.heroes = state.project.heroes.filter(h => h.id !== hero.id);
      state.selectedHeroId = filteredHeroes(true)[0]?.id || state.project.heroes[0]?.id || null;
      syncCompareSelection();
      renderAll();
    }

    async function renameHero() {
      const hero = getHeroById(state.selectedHeroId);
      if (!hero) return;
      const name = prompt('Rename hero:', hero['Hero Name'] || '');
      if (name === null) return;
      hero['Hero Name'] = name;
      renderAll();
      queueAutosave();
    }

    async function moveHeroToAge() {
      const hero = getHeroById(state.selectedHeroId);
      if (!hero) return;
      const target = prompt('Move hero to age:', hero.Age || state.selectedAge || '');
      if (!target) return;
      if (!state.project.ages.includes(target)) state.project.ages.push(target);
      hero.Age = target;
      state.selectedAge = target;
      renderAll();
      queueAutosave();
    }

    async function createAge() {
      const name = prompt('New age name:', 'New Age');
      if (!name) return;
      const data = await api('/api/age', { method: 'POST', body: JSON.stringify({ name }) });
      state.project.ages = data.ages;
      state.selectedAge = name;
      renderAll();
    }

    async function renameAge() {
      if (!state.selectedAge) return;
      const name = prompt('Rename age:', state.selectedAge);
      if (!name) return;
      const old = state.selectedAge;
      const data = await api(`/api/age/${encodeURIComponent(old)}`, { method: 'PUT', body: JSON.stringify({ name }) });
      state.project.ages = data.ages;
      state.project.heroes.forEach(hero => { if (hero.Age === old) hero.Age = name; });
      state.selectedAge = name;
      renderAll();
    }

    async function deleteAge() {
      if (!state.selectedAge || !confirm(`Delete age ${state.selectedAge}? Heroes will be moved.`)) return;
      const old = state.selectedAge;
      const data = await api(`/api/age/${encodeURIComponent(old)}`, { method: 'DELETE' });
      state.project.ages = data.ages;
      const fallback = state.project.ages[0];
      state.project.heroes.forEach(hero => { if (hero.Age === old) hero.Age = fallback; });
      state.selectedAge = fallback;
      renderAll();
      queueAutosave();
    }

    async function addField() {
      const name = prompt('Field name:', 'New Field');
      if (!name) return;
      const defaultCat = (state.project?.categories && state.project.categories[0]) || 'Custom';
      const category = prompt('Category:', defaultCat) || defaultCat;
      const type = prompt('Type (text, textarea, image, images):', 'text') || 'text';
      const data = await api('/api/field', { method: 'POST', body: JSON.stringify({ name, category, type }) });
      state.project.fields = data.fields;
      state.project.heroes.forEach(hero => { if (!(name in hero)) hero[name] = type === 'images' ? [] : ''; });
      renderAll();
      queueAutosave();
    }

    async function renameField(fieldName, maybeName, maybeCategory, maybeType) {
      const field = getFieldDefs().find(f => f.name === fieldName);
      if (!field) return;
      const name = maybeName ?? field.name;
      const category = maybeCategory ?? field.category;
      const type = maybeType ?? field.type;
      const data = await api(`/api/field/${encodeURIComponent(fieldName)}`, { method: 'PUT', body: JSON.stringify({ name, category, type }) });
      if (name !== fieldName) {
        state.project.heroes.forEach(hero => {
          hero[name] = hero[fieldName];
          delete hero[fieldName];
        });
      }
      state.project.fields = data.fields;
      renderAll();
      queueAutosave();
    }

    async function deleteField(fieldName) {
      if (!confirm(`Delete field ${fieldName}?`)) return;
      const data = await api(`/api/field/${encodeURIComponent(fieldName)}`, { method: 'DELETE' });
      state.project.fields = data.fields;
      state.project.heroes.forEach(hero => delete hero[fieldName]);
      renderAll();
      queueAutosave();
    }

    function currentHeroIndex() {
      const heroes = filteredHeroes(true);
      return heroes.findIndex(hero => hero.id === state.selectedHeroId);
    }

    function prevHero() {
      const heroes = filteredHeroes(true);
      const idx = currentHeroIndex();
      if (idx > 0) state.selectedHeroId = heroes[idx - 1].id;
      renderAll();
    }

    function nextHero() {
      const heroes = filteredHeroes(true);
      const idx = currentHeroIndex();
      if (idx >= 0 && idx < heroes.length - 1) state.selectedHeroId = heroes[idx + 1].id;
      renderAll();
    }

    function toggleCompare(id) {
      if (state.compareIds.includes(id)) state.compareIds = state.compareIds.filter(x => x !== id);
      else state.compareIds.push(id);
      renderComparePicker();
      if (state.mode === 'compare') runCompare();
    }

    function toggleCompareCurrent() {
      if (state.selectedHeroId) toggleCompare(state.selectedHeroId);
    }

    async function uploadPortrait(event) {
      const file = event.target.files[0];
      if (!file || !state.selectedHeroId) return;
      const form = new FormData();
      form.append('portrait', file);
      const hero = await api(`/api/hero/${state.selectedHeroId}/upload`, { method: 'POST', body: form });
      replaceHero(hero);
      renderAll();
    }

    async function uploadExtraImages(event) {
      const files = [...event.target.files];
      if (!files.length || !state.selectedHeroId) return;
      const form = new FormData();
      files.forEach(file => form.append('extra_images', file));
      const hero = await api(`/api/hero/${state.selectedHeroId}/upload`, { method: 'POST', body: form });
      replaceHero(hero);
      renderAll();
    }

    function replaceHero(hero) {
      const index = state.project.heroes.findIndex(h => h.id === hero.id);
      if (index >= 0) state.project.heroes[index] = hero;
      else state.project.heroes.push(hero);
    }

    async function promptProjectPath() {
      const current = state.project?.project_dir || '';
      const folder = prompt('Project folder path:', current);
      if (!folder) return;
      await api('/api/project/path', { method: 'POST', body: JSON.stringify({ project_dir: folder }) });
      await api('/api/import-existing', { method: 'POST', body: JSON.stringify({ project_dir: folder }) });
      await loadProject();
      setStatus('Project loaded');
    }

    function downloadWorkbook() {
      window.location.href = '/download/workbook';
    }

    function scrollToCategory(category) {
      const el = document.getElementById(`cat-${slug(category)}`);
      if (el) el.scrollIntoView({ behavior: 'smooth', block: 'start' });
    }

    function slug(value) {
      return String(value || '').toLowerCase().replace(/[^a-z0-9]+/g, '-').replace(/(^-|-$)/g, '');
    }

    function escapeHtml(value) {
      return String(value ?? '').replace(/[&<>"']/g, ch => ({ '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;' }[ch]));
    }

    function escapeAttr(value) {
      return escapeHtml(value).replace(/`/g, '&#96;');
    }

    function js(value) {
      return JSON.stringify(value).replace(/"/g, '&quot;');
    }

    window.addEventListener('beforeunload', () => {
      if (state.project) navigator.sendBeacon('/api/save', new Blob([JSON.stringify(state.project)], { type: 'application/json' }));
    });

    init();
  </script>
</body>
</html>
"""

# API Endpoints
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return HTMLResponse(INDEX_HTML)


@app.get("/api/project")
async def api_project():
    state = current_state()
    for hero in state["heroes"]:
        hero["_type"] = hero_display_type(hero)
    return state


@app.post("/api/project/path")
async def api_set_project_path(request: Request):
    payload = await request.json()
    target = Path(payload.get("project_dir", "")).expanduser().resolve()
    ensure_project_structure(target)
    save_config_legacy(target)
    ensure_workbook(target)
    return {"ok": True, "project_dir": str(target)}


@app.post("/api/save")
async def api_save(request: Request):
    state = await request.json()
    project_dir = Path(state.get("project_dir") or current_project_dir()).expanduser().resolve()
    write_state(project_dir, state)
    return {"ok": True, "saved_to": str(project_file(project_dir))}


@app.post("/api/import-existing")
async def api_import_existing(request: Request):
    payload = await request.json()
    target = Path(payload.get("project_dir", "")).expanduser().resolve()
    ensure_project_structure(target)
    save_config_legacy(target)
    if project_file(target).exists():
        state = import_existing_workbook(target)
    else:
        ensure_workbook(target)
        state = read_state(target)
    return state


@app.post("/api/hero")
async def api_create_hero(request: Request):
    state = current_state()
    payload = await request.json()
    ages = state["ages"] or ["Unassigned"]
    kingdoms = state.get("kingdoms", ["Humans"])
    hero = {field["name"]: "" for field in state["fields"] if field["name"] not in {"ID", "Extra Images"}}
    hero.update({
        "id": str(uuid.uuid4()),
        "Age": payload.get("Age") or ages[0],
        "Kingdom": payload.get("Kingdom") or kingdoms[0],
        "Hero Name": payload.get("Hero Name") or "New Hero",
        "Portrait": "",
        "Extra Images": [],
    })
    state["heroes"].append(hero)
    write_state(current_project_dir(), state)
    return hero


@app.put("/api/hero/{hero_id}")
async def api_update_hero(hero_id: str, request: Request):
    state = current_state()
    payload = await request.json()
    for hero in state["heroes"]:
        if hero["id"] == hero_id:
            hero.update(payload)
            write_state(current_project_dir(), state)
            return hero
    raise HTTPException(status_code=404, detail="Hero not found")


@app.post("/api/hero/{hero_id}/duplicate")
async def api_duplicate_hero(hero_id: str):
    state = current_state()
    for hero in state["heroes"]:
        if hero["id"] == hero_id:
            clone = json.loads(json.dumps(hero))
            clone["id"] = str(uuid.uuid4())
            clone["Hero Name"] = f"{hero.get('Hero Name', 'Hero')} Copy"
            state["heroes"].append(clone)
            write_state(current_project_dir(), state)
            return clone
    raise HTTPException(status_code=404, detail="Hero not found")


@app.delete("/api/hero/{hero_id}")
async def api_delete_hero(hero_id: str):
    state = current_state()
    before = len(state["heroes"])
    state["heroes"] = [hero for hero in state["heroes"] if hero["id"] != hero_id]
    if len(state["heroes"]) == before:
        raise HTTPException(status_code=404, detail="Hero not found")
    write_state(current_project_dir(), state)
    return {"ok": True}


@app.post("/api/hero/{hero_id}/upload")
async def api_upload_hero_image(hero_id: str, portrait: Optional[UploadFile] = File(None), extra_images: List[UploadFile] = File([])):
    state = current_state()
    hero = next((h for h in state["heroes"] if h["id"] == hero_id), None)
    if not hero:
        raise HTTPException(status_code=404, detail="Hero not found")

    project_dir = current_project_dir()
    
    if portrait:
        hero["Portrait"] = await save_image_async(portrait, project_dir)
    for image in extra_images:
        saved = await save_image_async(image, project_dir)
        if saved:
            hero.setdefault("Extra Images", []).append(saved)

    write_state(project_dir, state)
    return hero


async def save_image_async(file_storage: UploadFile, project_dir: Path):
    if not file_storage or not file_storage.filename:
        return ""
    ext = Path(file_storage.filename).suffix.lower() or ".png"
    safe_name = f"{uuid.uuid4().hex}{ext}"
    output = project_dir / "images" / safe_name
    with open(output, "wb") as f:
        content = await file_storage.read()
        f.write(content)
    return safe_name


@app.get("/images/{filename}")
async def serve_image(filename: str):
    return FileResponse(current_project_dir() / "images" / filename)


@app.get("/download/workbook")
async def download_workbook():
    project_dir = current_project_dir()
    ensure_workbook(project_dir)
    return FileResponse(project_dir / "Heroes.xlsx", filename="Heroes.xlsx")


@app.post("/api/age")
async def api_create_age(request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or "New Age"
    if name not in state["ages"]:
        state["ages"].append(name)
    write_state(current_project_dir(), state)
    return {"ages": state["ages"]}


@app.put("/api/age/{age_name}")
async def api_rename_age(age_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or age_name
    old = age_name
    state["ages"] = [name if a == old else a for a in state["ages"]]
    for hero in state["heroes"]:
        if hero["Age"] == old:
            hero["Age"] = name
    write_state(current_project_dir(), state)
    return {"ages": state["ages"]}


@app.delete("/api/age/{age_name}")
async def api_delete_age(age_name: str):
    state = current_state()
    old = age_name
    state["ages"] = [a for a in state["ages"] if a != old]
    fallback = state["ages"][0] if state["ages"] else "Unassigned"
    for hero in state["heroes"]:
        if hero["Age"] == old:
            hero["Age"] = fallback
    write_state(current_project_dir(), state)
    return {"ages": state["ages"]}


@app.post("/api/category")
async def api_create_category(request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or "New Category"
    if name not in state["categories"]:
        state["categories"].append(name)
    write_state(current_project_dir(), state)
    return {"categories": state["categories"]}


@app.put("/api/category/{category_name}")
async def api_rename_category(category_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or category_name
    old = category_name
    state["categories"] = [name if c == old else c for c in state["categories"]]
    for field in state["fields"]:
        if field.get("category") == old:
            field["category"] = name
    write_state(current_project_dir(), state)
    return {"categories": state["categories"], "fields": state["fields"]}


@app.delete("/api/category/{category_name}")
async def api_delete_category(category_name: str):
    state = current_state()
    old = category_name
    state["categories"] = [c for c in state["categories"] if c != old]
    fallback = "Custom"
    if fallback not in state["categories"]:
        state["categories"].append(fallback)
    for field in state["fields"]:
        if field.get("category") == old:
            field["category"] = fallback
    write_state(current_project_dir(), state)
    return {"categories": state["categories"], "fields": state["fields"]}


@app.post("/api/role")
async def api_create_role(request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or "New Role"
    if name not in state["roles"]:
        state["roles"].append(name)
    write_state(current_project_dir(), state)
    return {"roles": state["roles"]}


@app.put("/api/role/{role_name}")
async def api_rename_role(role_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or role_name
    old = role_name
    state["roles"] = [name if r == old else r for r in state["roles"]]
    for hero in state["heroes"]:
        if hero.get("Role") == old:
            hero["Role"] = name
    write_state(current_project_dir(), state)
    return {"roles": state["roles"]}


@app.delete("/api/role/{role_name}")
async def api_delete_role(role_name: str):
    state = current_state()
    old = role_name
    state["roles"] = [r for r in state["roles"] if r != old]
    write_state(current_project_dir(), state)
    return {"roles": state["roles"]}


@app.post("/api/god")
async def api_create_god(request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or "New God"
    if name not in state["gods"]:
        state["gods"].append(name)
    write_state(current_project_dir(), state)
    return {"gods": state["gods"]}


@app.put("/api/god/{god_name}")
async def api_rename_god(god_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or god_name
    old = god_name
    state["gods"] = [name if g == old else g for g in state["gods"]]
    for hero in state["heroes"]:
        if hero.get("God") == old:
            hero["God"] = name
    write_state(current_project_dir(), state)
    return {"gods": state["gods"]}


@app.delete("/api/god/{god_name}")
async def api_delete_god(god_name: str):
    state = current_state()
    old = god_name
    state["gods"] = [g for g in state["gods"] if g != old]
    write_state(current_project_dir(), state)
    return {"gods": state["gods"]}


@app.post("/api/kingdom")
async def api_create_kingdom(request: Request):
    state = current_state()
    payload = await request.json()
    name = normalize_text(payload.get("name")) or "New Kingdom"
    if name not in state["kingdoms"]:
        state["kingdoms"].append(name)
    # Also add to kingdoms_data if present
    if "kingdoms_data" not in state:
        state["kingdoms_data"] = {}
    state["kingdoms_data"][name] = {
        "description": payload.get("description", ""),
        "forefathers": payload.get("forefathers", ""),
        "established": payload.get("established", ""),
        "races": payload.get("races", ""),
        "history": payload.get("history", ""),
    }
    write_state(current_project_dir(), state)
    return {"kingdoms": state["kingdoms"], "kingdoms_data": state.get("kingdoms_data", {})}


@app.put("/api/kingdom/{kingdom_name}")
async def api_update_kingdom(kingdom_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    old = kingdom_name
    new_name = normalize_text(payload.get("name")) or old
    
    # Update name in list
    state["kingdoms"] = [new_name if k == old else k for k in state["kingdoms"]]
    
    # Update hero references
    for hero in state["heroes"]:
        if hero.get("Kingdom") == old:
            hero["Kingdom"] = new_name
    
    # Update kingdom data
    if "kingdoms_data" not in state:
        state["kingdoms_data"] = {}
    if old in state["kingdoms_data"]:
        state["kingdoms_data"][new_name] = state["kingdoms_data"].pop(old)
    
    # Update fields from payload
    state["kingdoms_data"][new_name] = {
        "description": payload.get("description", state["kingdoms_data"].get(new_name, {}).get("description", "")),
        "forefathers": payload.get("forefathers", state["kingdoms_data"].get(new_name, {}).get("forefathers", "")),
        "established": payload.get("established", state["kingdoms_data"].get(new_name, {}).get("established", "")),
        "races": payload.get("races", state["kingdoms_data"].get(new_name, {}).get("races", "")),
        "history": payload.get("history", state["kingdoms_data"].get(new_name, {}).get("history", "")),
    }
    
    write_state(current_project_dir(), state)
    return {"kingdoms": state["kingdoms"], "kingdoms_data": state["kingdoms_data"]}


@app.delete("/api/kingdom/{kingdom_name}")
async def api_delete_kingdom(kingdom_name: str):
    state = current_state()
    old = kingdom_name
    state["kingdoms"] = [k for k in state["kingdoms"] if k != old]
    fallback = state["kingdoms"][0] if state["kingdoms"] else "Humans"
    for hero in state["heroes"]:
        if hero.get("Kingdom") == old:
            hero["Kingdom"] = fallback
    if "kingdoms_data" in state and old in state["kingdoms_data"]:
        del state["kingdoms_data"][old]
    write_state(current_project_dir(), state)
    return {"kingdoms": state["kingdoms"], "kingdoms_data": state.get("kingdoms_data", {})}


@app.post("/api/compare")
async def api_compare(request: Request):
    payload = await request.json()
    hero_ids = payload.get("hero_ids", [])
    state = current_state()
    heroes = [h for h in state["heroes"] if h["id"] in hero_ids]

    if len(heroes) < 2:
        return {"heroes": heroes, "rows": []}

    all_fields = [f["name"] for f in state["fields"] if f["name"] not in {"ID", "Portrait", "Extra Images"}]
    rows = []
    for field in all_fields:
        values = [h.get(field, "") for h in heroes]
        first = values[0]
        different = any(v != first for v in values)
        rows.append({"field": field, "values": values, "different": different})

    return {"heroes": heroes, "rows": rows}


@app.post("/api/field")
async def api_add_field(request: Request):
    state = current_state()
    payload = await request.json()
    name = payload.get("name") or "New Field"
    category = payload.get("category") or "Custom"
    type_ = payload.get("type") or "text"
    field = {"name": name, "category": category, "type": type_}
    state["fields"].append(field)
    for hero in state["heroes"]:
        if name not in hero:
            hero[name] = [] if type_ == "images" else ""
    write_state(current_project_dir(), state)
    return {"fields": state["fields"]}


@app.put("/api/field/{field_name}")
async def api_rename_field(field_name: str, request: Request):
    state = current_state()
    payload = await request.json()
    name = payload.get("name") or field_name
    category = payload.get("category") or "Custom"
    type_ = payload.get("type") or "text"
    for field in state["fields"]:
        if field["name"] == field_name:
            if name != field_name:
                for hero in state["heroes"]:
                    hero[name] = hero[field_name]
                    del hero[field_name]
            field["name"] = name
            field["category"] = category
            field["type"] = type_
            break
    write_state(current_project_dir(), state)
    return {"fields": state["fields"]}


@app.delete("/api/field/{field_name}")
async def api_delete_field(field_name: str):
    state = current_state()
    state["fields"] = [f for f in state["fields"] if f["name"] != field_name]
    for hero in state["heroes"]:
        if field_name in hero:
            del hero[field_name]
    write_state(current_project_dir(), state)
    return {"fields": state["fields"]}


# Design Documents API
@app.get("/api/docs/list")
async def api_list_docs():
    design_dir = APP_ROOT / "design"
    design_dir.mkdir(exist_ok=True)
    files = sorted([f.name for f in design_dir.iterdir() if f.suffix.lower() == ".md"])
    return [{"filename": f} for f in files]


@app.get("/api/docs/read")
async def api_read_doc(filename: str):
    design_dir = APP_ROOT / "design"
    file_path = design_dir / filename
    # Security: ensure the file is within the design directory
    try:
        file_path.resolve().relative_to(design_dir.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid filename")
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    content = file_path.read_text(encoding="utf-8")
    return {"filename": filename, "content": content}


@app.post("/api/docs/save")
async def api_save_doc(request: Request):
    payload = await request.json()
    filename = payload.get("filename", "")
    content = payload.get("content", "")
    message = payload.get("message", f"Update {filename}")

    if not filename or not filename.endswith(".md"):
        raise HTTPException(status_code=400, detail="Filename must end with .md")

    design_dir = APP_ROOT / "design"
    design_dir.mkdir(exist_ok=True)
    file_path = design_dir / filename

    # Security check
    try:
        file_path.resolve().relative_to(design_dir.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid filename")

    # Write the file
    file_path.write_text(content, encoding="utf-8")

    # Git commit
    try:
        import subprocess
        subprocess.run(
            ["git", "add", str(file_path)],
            cwd=str(APP_ROOT),
            check=True,
            capture_output=True,
        )
        subprocess.run(
            ["git", "commit", "-m", message],
            cwd=str(APP_ROOT),
            check=False,
            capture_output=True,
        )
    except Exception:
        pass  # Git may fail if not configured, that's okay

    return {"ok": True, "filename": filename}


@app.get("/api/docs/history")
async def api_doc_history(filename: str):
    """Get git commit history for a specific document."""
    design_dir = APP_ROOT / "design"
    file_path = design_dir / filename
    
    # Security check
    try:
        file_path.resolve().relative_to(design_dir.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    try:
        import subprocess
        result = subprocess.run(
            ["git", "log", "--follow", "--format=%H|%ai|%s", "--", str(file_path)],
            cwd=str(APP_ROOT),
            capture_output=True,
            text=True,
        )
        history = []
        if result.returncode == 0 and result.stdout:
            for line in result.stdout.strip().split("\n"):
                if "|" in line:
                    parts = line.split("|", 2)
                    if len(parts) == 3:
                        hash_, date, message = parts
                        history.append({
                            "hash": hash_[:7],
                            "date": date.split(" ")[0] if " " in date else date,
                            "message": message,
                        })
        return {"filename": filename, "history": history}
    except Exception as e:
        return {"filename": filename, "history": [], "error": str(e)}


@app.get("/api/docs/version")
async def api_doc_version(filename: str, commit: str):
    """Get a specific version of a document from git history."""
    design_dir = APP_ROOT / "design"
    file_path = design_dir / filename
    
    # Security check
    try:
        file_path.resolve().relative_to(design_dir.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    try:
        import subprocess
        result = subprocess.run(
            ["git", "show", f"{commit}:{file_path.relative_to(APP_ROOT)}"],
            cwd=str(APP_ROOT),
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            return {"filename": filename, "commit": commit, "content": result.stdout}
        else:
            return {"filename": filename, "commit": commit, "content": "", "error": result.stderr}
    except Exception as e:
        return {"filename": filename, "commit": commit, "content": "", "error": str(e)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)